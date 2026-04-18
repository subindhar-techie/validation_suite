import os
import re
import traceback
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Import from same package using absolute imports
from .file_parsers import *
from .qr_processor import *
from .excel_generator import *
from .jio_validator import SCMReader, validate_jio_label, validate_outer_label_5000

# Import from utils using absolute path
import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from utils.helpers import *

def debug_pcom_content(pcom_path, key_patterns):
    """Debug function to see what's actually in the PCOM file"""
    print(f"\n🔍 DEBUG PCOM FILE: {pcom_path}")
    try:
        with open(pcom_path, 'r', encoding='utf-8', errors='ignore') as file:
            content = file.read()
            lines = content.split('\n')
            
        print("Relevant lines in PCOM file:")
        for i, line in enumerate(lines):
            line_lower = line.lower()
            # Check if line contains any of our target patterns
            if '.define' in line_lower:
                for var in ['%imsi', '%acc', '%puk', '%isc', '%iccid', '%adm', '%home_imsi', '%home_acc', '%dpuk1_card', '%dpuk2_card', '%adm1_card', '%iccid_card']:
                    if var in line_lower:
                        print(f"Line {i}: {line.strip()}")
                        break
                    
    except Exception as e:
        print(f"Error reading PCOM file for debug: {e}")

def extract_from_pcom_enhanced(file_path, patterns):
    """
    Enhanced PCOM extraction - searches entire file for patterns
    """
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            content = file.read()
        
        # Try multiple patterns
        if isinstance(patterns, str):
            patterns = [patterns]
        
        for pattern in patterns:
            match = re.search(pattern, content, re.IGNORECASE)
            if match:
                print(f"✅ PCOM pattern matched: {pattern} -> {match.group(1)}")
                return match.group(1)
        
        print(f"❌ No PCOM pattern matched from: {patterns}")
        return None
        
    except Exception as e:
        print(f"Error reading PCOM file {file_path}: {e}")
        return None

def _search_patterns(text, patterns):
    """Search text using multiple patterns"""
    if isinstance(patterns, str):
        patterns = [patterns]
    
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            print(f"Pattern matched: {pattern} -> {match.group(1)}")
            return match.group(1)
    
    print(f"No pattern matched from: {patterns}")
    return None

# ============================================================
# ENHANCED MACHINE LOG PARSING (Based on Airtel code)
# ============================================================

def extract_value_enhanced(line, command):
    """
    Enhanced value extraction - similar to Airtel approach
    """
    line = line.upper().strip()
    if command in line:
        parts = line.split(command)
        if len(parts) > 1:
            value = parts[1].split('SW9000')[0] if 'SW9000' in parts[1] else parts[1]
            # Clean up - allow hex characters
            value = re.sub(r'[^0-9A-F]', '', value)
            return value
    return None

def parse_machine_log_enhanced(filepath):
    """
    Enhanced machine log parsing using Airtel's approach
    """
    print("="*80)
    print("🚀 ENHANCED MACHINE LOG PARSING")
    print("="*80)
    
    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()
        
        extracted = {}
        
        for line_num, line in enumerate(lines, 1):
            line_upper = line.upper().strip()
            
            # ========== ICCID (2FE2) ==========
            if '2FE2' in line_upper and ('00A40000022FE2' in line_upper or 'SELECT 2FE2' in line_upper):
                # Look ahead for ICCID data
                for i in range(min(5, len(lines) - line_num)):
                    data_line = lines[line_num + i].upper().strip()
                    if '00D600000A' in data_line:
                        val = extract_value_enhanced(data_line, '00D600000A')
                        if val and len(val) >= 20:
                            extracted['ICCID_CARD (2FE2)'] = val[:20]
                            print(f"✅ Line {line_num}: ICCID parsed: {val[:20]}")
                        break
            
            # ========== IMSI (6F07) ==========
            elif '6F07' in line_upper and ('00A40000026F07' in line_upper or 'SELECT 6F07' in line_upper):
                for i in range(min(5, len(lines) - line_num)):
                    data_line = lines[line_num + i].upper().strip()
                    if '00D6000009' in data_line:
                        val = extract_value_enhanced(data_line, '00D6000009')
                        if val and len(val) >= 18:
                            # Convert hex to decimal digits (like Airtel)
                            decimal_digits = []
                            for j in range(0, len(val), 2):
                                if j + 2 <= len(val):
                                    hex_byte = val[j:j+2]
                                    try:
                                        byte_value = int(hex_byte, 16)
                                        high_digit = (byte_value >> 4) & 0x0F
                                        low_digit = byte_value & 0x0F
                                        
                                        if 0 <= high_digit <= 9:
                                            decimal_digits.append(str(high_digit))
                                        if 0 <= low_digit <= 9:
                                            decimal_digits.append(str(low_digit))
                                    except:
                                        continue
                            
                            if decimal_digits and len(decimal_digits) >= 18:
                                imsi_18_digit = ''.join(decimal_digits)[:18]
                                extracted['HOME_IMSI (6F07)'] = imsi_18_digit
                                print(f"✅ Line {line_num}: HOME_IMSI parsed: {imsi_18_digit}")
                            elif decimal_digits:
                                imsi = ''.join(decimal_digits)
                                extracted['HOME_IMSI (6F07)'] = imsi
                                print(f"✅ Line {line_num}: HOME_IMSI parsed: {imsi}")
                        break
            
            # ========== PSK/DEK1 (6F2B) ==========
            elif '6F2B' in line_upper and ('00A40000026F2B' in line_upper or 'SELECT 6F2B' in line_upper):
                for i in range(min(5, len(lines) - line_num)):
                    data_line = lines[line_num + i].upper().strip()
                    if '00D600002AFE85410110' in data_line and 'FE80410210' in data_line:
                        parts = data_line.split('00D600002AFE85410110')
                        if len(parts) > 1:
                            tail = parts[1]
                            if 'FE80410210' in tail:
                                psk = tail.split('FE80410210')[0][:32]
                                dek1 = tail.split('FE80410210')[1][:32]
                                extracted['PSK (6F2B)'] = psk
                                extracted['DEK1 (6F2B)'] = dek1
                                print(f"✅ Line {line_num}: PSK parsed: {psk[:16]}...")
                                print(f"✅ Line {line_num}: DEK1 parsed: {dek1[:16]}...")
                        break
            
            # ========== DPUK1 (6F01) ==========
            elif '6F01' in line_upper and ('00A40000026F01' in line_upper or 'SELECT 6F01' in line_upper):
                for i in range(min(5, len(lines) - line_num)):
                    data_line = lines[line_num + i].upper().strip()
                    if '00D6000015F00A0A' in data_line:
                        val = extract_value_enhanced(data_line, '00D6000015F00A0A')
                        if val and len(val) >= 16:
                            if 'FFFFFFFF0A0A' in val:
                                parts = val.split('FFFFFFFF0A0A')
                                if len(parts) > 1 and len(parts[1]) >= 16:
                                    extracted['DPUK1_CARD (6F01)'] = parts[1][:16]
                                    print(f"✅ Line {line_num}: DPUK1 parsed: {parts[1][:16]}")
                            else:
                                extracted['DPUK1_CARD (6F01)'] = val[:16]
                                print(f"✅ Line {line_num}: DPUK1 parsed: {val[:16]}")
                        break
            
            # ========== DPUK2 (6F81) ==========
            elif '6F81' in line_upper and ('00A40000026F81' in line_upper or 'SELECT 6F81' in line_upper):
                for i in range(min(5, len(lines) - line_num)):
                    data_line = lines[line_num + i].upper().strip()
                    if '00D6000015E00A0A' in data_line:
                        val = extract_value_enhanced(data_line, '00D6000015E00A0A')
                        if val and len(val) >= 16:
                            if 'FFFFFFFF0A0A' in val:
                                parts = val.split('FFFFFFFF0A0A')
                                if len(parts) > 1 and len(parts[1]) >= 16:
                                    extracted['DPUK2_CARD (6F81)'] = parts[1][:16]
                                    print(f"✅ Line {line_num}: DPUK2 parsed: {parts[1][:16]}")
                            else:
                                extracted['DPUK2_CARD (6F81)'] = val[:16]
                                print(f"✅ Line {line_num}: DPUK2 parsed: {val[:16]}")
                        break
            
            # ========== ADM (6F0A) ==========
            elif '6F0A' in line_upper and ('00A40000026F0A' in line_upper or 'SELECT 6F0A' in line_upper):
                for i in range(min(5, len(lines) - line_num)):
                    data_line = lines[line_num + i].upper().strip()
                    if '00D600000B800A0A' in data_line:
                        val = extract_value_enhanced(data_line, '00D600000B800A0A')
                        if val and len(val) >= 16:
                            extracted['ADM (6F0A)'] = val[:16]
                            print(f"✅ Line {line_num}: ADM parsed: {val[:16]}")
                        break
            
            # ========== HOME_ACC (6F78) ==========
            elif '6F78' in line_upper and ('00A40000026F78' in line_upper or 'SELECT 6F78' in line_upper):
                for i in range(min(5, len(lines) - line_num)):
                    data_line = lines[line_num + i].upper().strip()
                    if '00D6000002' in data_line:
                        val = extract_value_enhanced(data_line, '00D6000002')
                        if val and len(val) >= 4:
                            extracted['HOME_ACC (6F78)'] = val[:4]
                            print(f"✅ Line {line_num}: HOME_ACC parsed: {val[:4]}")
                        break
            
            # ========== GLOBAL_IMSI (3031) ==========
            elif '3031' in line_upper and ('00A40000023031' in line_upper or 'SELECT 3031' in line_upper):
                for i in range(min(5, len(lines) - line_num)):
                    data_line = lines[line_num + i].upper().strip()
                    if '00D6000009' in data_line:
                        val = extract_value_enhanced(data_line, '00D6000009')
                        if val and len(val) >= 18:
                            # Convert hex to decimal digits
                            decimal_digits = []
                            for j in range(0, len(val), 2):
                                if j + 2 <= len(val):
                                    hex_byte = val[j:j+2]
                                    try:
                                        byte_value = int(hex_byte, 16)
                                        high_digit = (byte_value >> 4) & 0x0F
                                        low_digit = byte_value & 0x0F
                                        
                                        if 0 <= high_digit <= 9:
                                            decimal_digits.append(str(high_digit))
                                        if 0 <= low_digit <= 9:
                                            decimal_digits.append(str(low_digit))
                                    except:
                                        continue
                            
                            if decimal_digits and len(decimal_digits) >= 18:
                                imsi_18_digit = ''.join(decimal_digits)[:18]
                                extracted['GLOBAL_IMSI (3031)'] = imsi_18_digit
                                print(f"✅ Line {line_num}: GLOBAL_IMSI parsed: {imsi_18_digit}")
                        break
            
            # ========== GLOBAL_ACC (3037) ==========
            elif '3037' in line_upper and ('00A40000023037' in line_upper or 'SELECT 3037' in line_upper):
                for i in range(min(5, len(lines) - line_num)):
                    data_line = lines[line_num + i].upper().strip()
                    if '00D60000120000000300000002FFFFFFFF' in data_line:
                        val = extract_value_enhanced(data_line, '00D60000120000000300000002FFFFFFFF')
                        if val and len(val) >= 8:
                            extracted['GLOBAL_ACC (3037)'] = val[:4]
                            extracted['HOME_ACC (3037)'] = val[4:8]
                            print(f"✅ Line {line_num}: GLOBAL_ACC parsed: {val[:4]}")
                            print(f"✅ Line {line_num}: HOME_ACC parsed: {val[4:8]}")
                        break
            
            # ========== KIC/KID Keys (6F22) ==========
            elif '6F22' in line_upper and ('00A40000026F22' in line_upper or 'SELECT 6F22' in line_upper):
                kic_kid_prefixes = {
                    'KIC1 (6F22)': '00DC01041BFE0150',
                    'KID1 (6F22)': '00DC02041BFE0151',
                    'KIK1 (6F22)': '00DC03041BFE0152',
                    'KIC2 (6F22)': '00DC04041BFE0250',
                    'KID2 (6F22)': '00DC05041BFE0251',
                    'KIK2 (6F22)': '00DC06041BFE0252',
                }
                
                for prefix_key, prefix in kic_kid_prefixes.items():
                    for i in range(min(10, len(lines) - line_num)):
                        data_line = lines[line_num + i].upper().strip()
                        if prefix in data_line:
                            val = extract_value_enhanced(data_line, prefix)
                            if val:
                                val = val.replace('FFFFFFFFFFFFFFFF', '')
                                if len(val) >= 32:
                                    extracted[prefix_key] = val[:32]
                                    print(f"✅ Line {line_num+i+1}: {prefix_key} parsed: {val[:16]}...")
                            break
            
            # ========== ASCII IMSI (6F02) ==========
            elif '6F02' in line_upper and ('00A40000026F02' in line_upper or 'SELECT 6F02' in line_upper):
                for i in range(min(5, len(lines) - line_num)):
                    data_line = lines[line_num + i].upper().strip()
                    if '00D600005F8031' in data_line:
                        parts = data_line.split('00D600005F8031')
                        if len(parts) > 1:
                            val = parts[1]
                            extracted['ASCII_IMSI (6F02)'] = val[:30]
                            print(f"✅ Line {line_num}: ASCII_IMSI (6F02) parsed: {val[:30]}")
                        break
            
            # ========== ASCII IMSI (6F04) ==========
            elif '6F04' in line_upper and ('00A40000026F04' in line_upper or 'SELECT 6F04' in line_upper):
                for i in range(min(5, len(lines) - line_num)):
                    data_line = lines[line_num + i].upper().strip()
                    if '00DC01047880357369703A' in data_line:
                        parts = data_line.split('00DC01047880357369703A')
                        if len(parts) > 1:
                            val = parts[1]
                            extracted['ASCII_IMSI (6F04)'] = val[:30]
                            print(f"✅ Line {line_num}: ASCII_IMSI (6F04) parsed: {val[:30]}")
                        break
        
        print(f"\n✅ Extracted {len(extracted)} fields from Machine Log")
        for field, value in extracted.items():
            print(f"   - {field}: {value}")
        
        return extracted
        
    except Exception as e:
        print(f"❌ Machine Log parsing error: {str(e)}")
        traceback.print_exc()
        return {}

def parse_machine_log_robust(filepath):
    """
    Robust machine log parsing with multiple strategies
    """
    print("\n" + "="*80)
    print("ROBUST MACHINE LOG PARSING")
    print("="*80)
    
    # Strategy 1: Enhanced parsing (Airtel style)
    results1 = parse_machine_log_enhanced(filepath)
    
    # Strategy 2: Pattern-based search (fallback)
    results2 = {}
    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        
        # Pattern search for specific fields
        patterns = [
            ('ICCID_CARD (2FE2)', r'00D600000A([0-9A-F]{20})'),
            ('HOME_IMSI (6F07)', r'00D6000009([0-9A-F]{18})'),
            ('PSK (6F2B)', r'FE85410110([0-9A-F]{32})'),
            ('DEK1 (6F2B)', r'FE80410210([0-9A-F]{32})'),
            ('DPUK1_CARD (6F01)', r'F00A0A([0-9A-F]{16})'),
            ('DPUK2_CARD (6F81)', r'E00A0A([0-9A-F]{16})'),
            ('ADM (6F0A)', r'800A0A([0-9A-F]{16})'),
            ('HOME_ACC (6F78)', r'00D6000002([0-9A-F]{4})'),
            ('GLOBAL_IMSI (3031)', r'00A40000023031[^\n]*?\n[^\n]*?00D6000009([0-9A-F]{18})'),
        ]
        
        for field, pattern in patterns:
            matches = re.findall(pattern, content, re.IGNORECASE)
            if matches:
                results2[field] = matches[0]
                print(f"✅ Pattern found {field}: {matches[0]}")
    
    except Exception as e:
        print(f"Pattern search error: {e}")
    
    # Merge results (prefer results1, fallback to results2)
    all_results = {}
    all_results.update(results2)  # Pattern results first
    all_results.update(results1)  # Enhanced results override
    
    print(f"\n📊 Total extracted: {len(all_results)} fields")
    return all_results

class ValidationEngine:
    def __init__(self):
        self.results = {}

def main(profile_type, filepath, pcom_path, scm_path, sim_oda_path, cnum_path=None, image_paths=None, circle_value=None, perso_script_path=None):
    
    print("Running with:", profile_type, filepath)

    # Initialize error collection
    validation_errors = []

    # Fields to skip per profile
    skip_map = {
        "MOB": ["INNER_LABEL_100"],
        "WBIOT": ["GLOBAL_IMSI (3031)", "GLOBAL_ACC (3037)", "HOME_ACC (3037)"],
        "NBIOT": ["GLOBAL_IMSI (3031)", "GLOBAL_ACC (3037)", "HOME_ACC (3037)", 
                  "ASCII_IMSI (6F02)", "ASCII_IMSI (6F04)"]
    }

    skip_fields = set(skip_map.get(profile_type, []))

    # Field definitions
    key_to_header = {
        "PSK (6F2B)": "PSK (6F2B)",
        "DEK1 (6F2B)": "DEK1 (6F2B)",
        "GLOBAL_IMSI (3031)": "GLOBAL_IMSI (3031)",
        "HOME_IMSI (6F07)": "HOME_IMSI (6F07)",
        "GLOBAL_ACC (3037)": "GLOBAL_ACC (3037)",
        "HOME_ACC (6F78)": "HOME_ACC (6F78)",
        "DPUK1_CARD (6F01)": "DPUK1_CARD (6F01)",
        "DPUK2_CARD (6F81)": "DPUK2_CARD (6F81)",
        "ADM (6F0A)": "ADM (6F0A)",
        "ICCID_CARD (2FE2)": "ICCID_CARD (2FE2)",
        "KIC1 (6F22)": "KIC1 (6F22)",
        "KID1 (6F22)": "KID1 (6F22)",
        "KIC2 (6F22)": "KIC2 (6F22)",
        "KID2 (6F22)": "KID2 (6F22)",
        "ASCII_IMSI (6F02)": "ASCII_IMSI (6F02)",
        "ASCII_IMSI (6F04)": "ASCII_IMSI (6F04)",
        "HOME_ACC (3037)": "HOME_ACC (3037)",
    }

    # Enhanced PCOM configuration with multiple search patterns
    if profile_type == 'NBIOT':
        pcom_config = {
            "HOME_IMSI (6F07)": [
                r"\.DEFINE\s+%IMSI\s+([0-9]+)",
                r"\.DEFINE\s+%HOME_IMSI\s+([0-9]+)",
                r"IMSI\s*=\s*([0-9]+)"
            ],
            "HOME_ACC (6F78)": [
                r"\.DEFINE\s+%ACC\s+([0-9]+)", 
                r"\.DEFINE\s+%HOME_ACC\s+([0-9]+)",
                r"ACC\s*=\s*([0-9]+)"
            ],
            "DPUK1_CARD (6F01)": [
                r"\.DEFINE\s+%PUK1\s+([0-9]+)",
                r"\.DEFINE\s+%DPUK1_CARD\s+([0-9]+)",
                r"PUK1\s*=\s*([0-9]+)"
            ],
            "DPUK2_CARD (6F81)": [
                r"\.DEFINE\s+%PUK2\s+([0-9]+)",
                r"\.DEFINE\s+%DPUK2_CARD\s+([0-9]+)",
                r"PUK2\s*=\s*([0-9]+)"
            ],
            "ADM (6F0A)": [
                r"\.DEFINE\s+%ISC1\s+([0-9A-Fa-f]+)",
                r"\.DEFINE\s+%ADM1_CARD\s+([0-9A-Fa-f]+)",
                r"ADM\s*=\s*([0-9A-Fa-f]+)"
            ],
            "ICCID_CARD (2FE2)": [
                r"\.DEFINE\s+%ICCID\s+([0-9]+)",
                r"\.DEFINE\s+%ICCID_CARD\s+([0-9]+)",
                r"ICCID\s*=\s*([0-9]+)"
            ]
        }
    else: 
        pcom_config = {
            "GLOBAL_IMSI (3031)": [
                r"\.DEFINE\s+%HOME_IMSI\s+([0-9]+)",
                r"IMSI\s*=\s*([0-9]+)"
            ],
            "HOME_IMSI (6F07)": [
                r"\.DEFINE\s+%HOME_IMSI\s+([0-9]+)",
                r"IMSI\s*=\s*([0-9]+)"
            ],
            "GLOBAL_ACC (3037)": [
                r"\.DEFINE\s+%HOME_ACC\s+([0-9]+)",
                r"ACC\s*=\s*([0-9]+)"
            ],
            "HOME_ACC (6F78)": [
                r"\.DEFINE\s+%HOME_ACC\s+([0-9]+)",
                r"ACC\s*=\s*([0-9]+)"
            ],
            "DPUK1_CARD (6F01)": [
                r"\.DEFINE\s+%PUK1\s+([0-9]+)",
                r"\.DEFINE\s+%DPUK1_CARD\s+([0-9]+)",
                r"PUK1\s*=\s*([0-9]+)"
            ],
            "DPUK2_CARD (6F81)": [
                r"\.DEFINE\s+%PUK2\s+([0-9]+)",
                r"\.DEFINE\s+%DPUK2_CARD\s+([0-9]+)",
                r"PUK2\s*=\s*([0-9]+)"
            ],
            "ADM (6F0A)": [
                r"\.DEFINE\s+%ISC1\s+([0-9A-Fa-f]+)",
                r"\.DEFINE\s+%ADM1_CARD\s+([0-9A-Fa-f]+)",
                r"ADM\s*=\s*([0-9A-Fa-f]+)"
            ],
            "ICCID_CARD (2FE2)": [
                r"\.DEFINE\s+%ICCID\s+([0-9]+)",
                r"\.DEFINE\s+%ICCID_CARD\s+([0-9]+)",
                r"ICCID\s*=\s*([0-9]+)"
            ],
            "ASCII_IMSI (6F02)": [
                r"\.DEFINE\s+%ASCII_IMSI\s+([0-9]+)",
                r"ASCII_IMSI\s*=\s*([0-9]+)"
            ],
            "ASCII_IMSI (6F04)": [
                r"\.DEFINE\s+%ASCII_IMSI\s+([0-9]+)",
                r"ASCII_IMSI\s*=\s*([0-9]+)"
            ],
            "HOME_ACC (3037)": [
                r"\.DEFINE\s+%HOME_ACC\s+([0-9]+)",
                r"ACC\s*=\s*([0-9]+)"
            ]
        }
    
    cnum_config = {
        "GLOBAL_IMSI (3031)": (16, 2, True),  # Special logic, only first value
        "HOME_IMSI (6F07)": (16, 2, True),    # Special logic, only first value
        "ASCII_IMSI (6F02)": (16, 2, True),   # Special logic, only first value
        "ASCII_IMSI (6F04)": (16, 2, True),   # Special logic, only first value
        "ICCID_CARD (2FE2)": (16, 4, False),  # Standard logic
        "DPUK1_CARD (6F01)": (16, 6, False),  # Standard logic
        "DPUK2_CARD (6F81)": (16, 8, False)   # Standard logic
    }

    scm_config = {
        "GLOBAL_IMSI (3031)": (2, 3),   # Column 4 (0-indexed)
        "HOME_IMSI (6F07)": (2, 3),
        "ICCID_CARD (2FE2)": (2, 2),    # Column 3
        "ASCII_IMSI (6F02)": (2, 3),
        "ASCII_IMSI (6F04)": (2, 3)
    }
    
    if profile_type == 'NBIOT':
        sim_oda_config = {
            "PSK (6F2B)": (126, r"SecurityKey\([^,]+, [^,]+, PskTls, ([^,]+),"),
            "DEK1 (6F2B)": (127, r"SecurityKey\([^,]+, [^,]+, Management, ([^,]+),"),
            "GLOBAL_IMSI (3031)": (130, r"Imsi\((\w+)\)"),
            "HOME_IMSI (6F07)": (130, r"Imsi\((\w+)\)"),
            "ICCID_CARD (2FE2)": (121, r"Iccid\(([^,]+),.*\)"),
            "KIC1 (6F22)": (122, r"SecurityKey\(.*, Encryption, (\w+)\)"),
            "KID1 (6F22)": (123, r"SecurityKey\(.*, Authentication, (\w+)\)"),
            "KIC2 (6F22)": (124, r"SecurityKey\(.*, Encryption, (\w+)\)"),
            "KID2 (6F22)": (125, r"SecurityKey\(.*, Authentication, (\w+)\)"),
            "ASCII_IMSI (6F02)": (130, r"Imsi\((\w+)\)"),
            "ASCII_IMSI (6F04)": (130, r"Imsi\((\w+)\)"),
        }
    else:
        sim_oda_config = {
            "PSK (6F2B)": (351, r"SecurityKey\([^,]+, [^,]+, PskTls, ([^,]+),"),
            "DEK1 (6F2B)": (352, r"SecurityKey\([^,]+, [^,]+, Management, ([^,]+),"),
            "GLOBAL_IMSI (3031)": (355, r"Imsi\((\w+)\)"),
            "HOME_IMSI (6F07)": (355, r"Imsi\((\w+)\)"),
            "ICCID_CARD (2FE2)": (346, r"Iccid\(([^,]+),.*\)"),
            "KIC1 (6F22)": (347, r"SecurityKey\(.*, Encryption, (\w+)\)"),
            "KID1 (6F22)": (348, r"SecurityKey\(.*, Authentication, (\w+)\)"),
            "KIC2 (6F22)": (349, r"SecurityKey\(.*, Encryption, (\w+)\)"),
            "KID2 (6F22)": (350, r"SecurityKey\(.*, Authentication, (\w+)\)"),
            "ASCII_IMSI (6F02)": (355, r"Imsi\((\w+)\)"),
            "ASCII_IMSI (6F04)": (355, r"Imsi\((\w+)\)")
        }

    # Extract values from all files
    file_values = {k: {"PCOM": None, "CNUM": None, "SCM": None, "SIM_ODA": None} 
                for k in key_to_header if k not in skip_fields}
    
    # ============================================================
    # ENHANCED PCOM EXTRACTION
    # ============================================================
    print("\n" + "="*80)
    print("EXTRACTING FROM PCOM FILE")
    print("="*80)
    
    for key, patterns in pcom_config.items():
        if key in skip_fields:
            continue
        
        print(f"\n🔍 Searching for {key} in PCOM...")
        extracted_value = extract_from_pcom_enhanced(pcom_path, patterns)
        file_values[key]["PCOM"] = extracted_value
        
        if extracted_value is None:
            error_msg = f"[{key}] PCOM extraction failed"
            print(f"❌ {error_msg}")
            validation_errors.append(error_msg)
        else:
            print(f"✅ [{key}] PCOM extraction successful: {extracted_value}")

    # ============================================================
    # ENHANCED MACHINE LOG PARSING
    # ============================================================
    print("\n" + "="*80)
    print("EXTRACTING FROM MACHINE LOG")
    print("="*80)
    
    results = parse_machine_log_robust(filepath)
    
    # Initialize all fields
    for key in key_to_header.keys():
        if key not in results:
            results[key] = "N/A"
    
    # ============================================================
    # EXTRACT FROM OTHER FILES
    # ============================================================
    
    # First handle KIC/KID keys using ordered match
    kic_matches = extract_multiple_keys(sim_oda_path, r"SecurityKey\(.*, Encryption, (\w+)\)")
    kid_matches = extract_multiple_keys(sim_oda_path, r"SecurityKey\(.*, Authentication, (\w+)\)")

    if "KIC1 (6F22)" in file_values:
        file_values["KIC1 (6F22)"]["SIM_ODA"] = kic_matches[0] if len(kic_matches) > 0 else None
    if "KIC2 (6F22)" in file_values:
        file_values["KIC2 (6F22)"]["SIM_ODA"] = kic_matches[1] if len(kic_matches) > 1 else None
    if "KID1 (6F22)" in file_values:
        file_values["KID1 (6F22)"]["SIM_ODA"] = kid_matches[0] if len(kid_matches) > 0 else None
    if "KID2 (6F22)" in file_values:
        file_values["KID2 (6F22)"]["SIM_ODA"] = kid_matches[1] if len(kid_matches) > 1 else None

    # Extract from CNUM
    for key, (line_num, col_idx, special_logic) in cnum_config.items():
        if key in skip_fields:
            continue
        if key in file_values:
            file_values[key]["CNUM"] = extract_from_cnum(cnum_path, line_num, col_idx, special_logic)

    # Extract from SCM
    for key, (line_num, col_idx) in scm_config.items():
        if key in skip_fields:
            continue
        if key in file_values:
            file_values[key]["SCM"] = extract_from_scm(scm_path, line_num, col_idx)

    # Extract from SIM ODA
    for key, (line_num, pattern) in sim_oda_config.items():
        if key not in file_values or key in ["KIC1 (6F22)", "KIC2 (6F22)", "KID1 (6F22)", "KID2 (6F22)"]:
            continue
        file_values[key]["SIM_ODA"] = extract_from_sim_oda(
            sim_oda_path, line_num, pattern, search_range=2, fallback=True
        )

    # ============================================================
    # VALIDATION RULES
    # ============================================================
    rules = {
        "PSK (6F2B)": ("NR", "NR", "NR", "from_value"),
        "DEK1 (6F2B)": ("NR", "NR", "NR", "from_value"),
        "GLOBAL_IMSI (3031)": ("from_value", "from_value", "from_value", "from_value"),
        "HOME_IMSI (6F07)": ("from_value", "from_value", "from_value", "from_value"),
        "GLOBAL_ACC (3037)": ("from_value", "NR", "NR", "NR"),
        "HOME_ACC (6F78)": ("from_value", "NR", "NR", "NR"),
        "DPUK1_CARD (6F01)": ("from_value", "from_value", "NR", "NR"),
        "DPUK2_CARD (6F81)": ("from_value", "from_value", "NR", "NR"),
        "ADM (6F0A)": ("from_value", "NR", "NR", "NR"),
        "ICCID_CARD (2FE2)": ("from_value", "from_value", "from_value", "from_value"),
        "KIC1 (6F22)": ("NR", "NR", "NR", "from_value"),
        "KID1 (6F22)": ("NR", "NR", "NR", "from_value"),
        "KIC2 (6F22)": ("NR", "NR", "NR", "from_value"),
        "KID2 (6F22)": ("NR", "NR", "NR", "from_value"),
        "ASCII_IMSI (6F02)": ("from_value", "from_value", "from_value", "from_value"),
        "ASCII_IMSI (6F04)": ("from_value", "from_value", "from_value", "from_value"),
        "HOME_ACC (3037)": ("from_value", "NR", "NR", "NR"),
    }

    # Excel Report Setup
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation Report"

    if image_paths:
        # Set starting row based on profile_type
        if profile_type == "NBIOT":
            start_row = 30
        elif profile_type == "WBIOT":
            start_row = 32
        else:
            start_row = 30  # Default start row

        start_col_letter = 'D'  # Column D

        for i, image_path in enumerate(image_paths):
            if image_path:  # If a valid image path is provided
                row = start_row + (i * 10)  # Increase row by 10 for each image
                cell = f'{start_col_letter}{row}'  # e.g., D32, D42, D52, etc.
                insert_image(ws, image_path, cell)  # Insert image at the specified cell

    # Styles
    styles = setup_excel_styles()

    # Headers
    headers = ["Field", "Machine Log", "PCOM", "CNUM", "SCM", "SIM ODA", 
               "PCOM Status", "CNUM Status", "SCM Status", "SIM_ODA Status", "Validation Status"]
    
    # Setup Excel headers and metadata with Final Verification Report Card values
    # Get folder path from machine log file
    folder_path = os.path.dirname(filepath) if filepath else ""
    
    setup_excel_headers(
        ws, styles, 
        operator_name="JIO",
        folder_path=folder_path,
        scm_path=scm_path,
        sim_oda_path=sim_oda_path,
        circle_value=circle_value,
        validation_errors=validation_errors,
        cnum_path=cnum_path,
        perso_script_path=perso_script_path
    )

    header_row = 16  # Row 18 will now contain your table headers
           
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = styles['dark_blue_fill']
        cell.font = styles['header_font']
        cell.border = styles['thick_border']
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Iterate over each key
    row = header_row + 1 

    mismatch_logs = []  

    for key in key_to_header.keys():
        if key in skip_fields:
            continue

        ml_val = results.get(key, "N/A")
        rule = rules.get(key, ("NR", "NR", "NR", "NR"))

        pcom_raw = file_values.get(key, {}).get("PCOM")
        cnum_raw = file_values.get(key, {}).get("CNUM")
        scm_raw = file_values.get(key, {}).get("SCM")
        sim_oda_raw = file_values.get(key, {}).get("SIM_ODA")

        pcom_disp = pcom_raw if rule[0] == "from_value" and pcom_raw else "Missing" if rule[0] == "from_value" else "NR"
        cnum_disp = cnum_raw if rule[1] == "from_value" and cnum_raw else "Missing" if rule[1] == "from_value" else "NR"
        scm_disp = scm_raw if rule[2] == "from_value" and scm_raw else "Missing" if rule[2] == "from_value" else "NR"
        sim_oda_disp = sim_oda_raw if rule[3] == "from_value" and sim_oda_raw else "Missing" if rule[3] == "from_value" else "NR"

        status = {"PCOM": "NR", "CNUM": "NR", "SCM": "NR", "SIM_ODA": "NR"}
        overall_valid = True

        def is_ascii_imsi_key(k): return "ASCII_IMSI" in k
        def is_imsi_key(k): return "IMSI" in k and "ASCII_IMSI" not in k
        def is_iccid_key(k): return k == "ICCID_CARD (2FE2)"
        def is_dpuk_key(k): return k in ["DPUK1_CARD (6F01)", "DPUK2_CARD (6F81)"]

        if rule[0] == "from_value":
            if pcom_raw:
                if len(pcom_raw) != len(ml_val):
                    error_msg = f"[{key}] PCOM length mismatch: ML length={len(ml_val)}, PCOM length={len(pcom_raw)}"
                    mismatch_logs.append(
                        f"[{key}] PCOM length mismatch: ML length={len(ml_val)}, PCOM length={len(pcom_raw)}, "
                        f"ML value={ml_val}, PCOM value={pcom_raw}"
                    )
                    validation_errors.append(error_msg)
                    status["PCOM"] = "❌ Fail"
                    overall_valid = False
                else:
                    if is_iccid_key(key):
                        norm_ml, norm = normalize_iccid(ml_val), normalize_iccid(pcom_raw)
                    elif is_ascii_imsi_key(key):
                        norm_ml, norm = normalize_ascii_imsi(ml_val), normalize_ascii_imsi(pcom_raw)
                    elif is_imsi_key(key):
                        norm_ml, norm = normalize_imsi(ml_val), normalize_imsi(pcom_raw)
                    else:
                        norm_ml, norm = ml_val, pcom_raw
                    if norm_ml == norm:
                        status["PCOM"] = "✅ Pass"
                    else:
                        status["PCOM"] = "❌ Fail"
                        error_msg = f"[{key}] PCOM mismatch: ML={norm_ml}, PCOM={norm}"
                        mismatch_logs.append(error_msg)
                        validation_errors.append(error_msg)
                        overall_valid = False
            else:
                status["PCOM"] = "❌ Fail"
                error_msg = f"[{key}] PCOM Missing"
                mismatch_logs.append(error_msg)
                validation_errors.append(error_msg)
                overall_valid = False

        # DPUK special handling
        if is_dpuk_key(key):
            try:
                decoded_ml = bytes.fromhex(ml_val).decode('ascii')
            except ValueError:
                decoded_ml = ""

            norm_cnum = re.sub(r'\D', '', cnum_raw or "")
            norm_pcom = re.sub(r'\D', '', pcom_raw or "")

            status = {
                "PCOM": "✅ Pass" if ml_val == norm_pcom else "❌ Fail" if pcom_raw else "Missing",
                "CNUM": "✅ Pass" if decoded_ml == norm_cnum else "❌ Fail" if cnum_raw else "Missing",
                "SCM": "NR",
                "SIM_ODA": "NR"
            }

            if status["PCOM"] != "✅ Pass":
                error_msg = f"[{key}] DPUK PCOM mismatch: ML={ml_val}, PCOM={norm_pcom}"
                mismatch_logs.append(error_msg)
                validation_errors.append(error_msg)
            if status["CNUM"] != "✅ Pass":
                error_msg = f"[{key}] DPUK CNUM mismatch: Decoded ML={decoded_ml}, CNUM={norm_cnum}"
                mismatch_logs.append(error_msg)
                validation_errors.append(error_msg)

            overall_valid = all(v == "✅ Pass" or v == "NR" for v in status.values())

            # Display values
            pcom_disp = pcom_raw or "Missing"
            cnum_disp = cnum_raw or "Missing"
            scm_disp = sim_oda_disp = "NR"
            validation_status = "✅ Pass" if overall_valid else "❌ Fail"
            data = [key, ml_val, pcom_disp, cnum_disp, scm_disp, sim_oda_disp,
                    status["PCOM"], status["CNUM"], status["SCM"], status["SIM_ODA"], validation_status]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = styles['thick_border']
                cell.number_format = '@'
                if "Pass" in str(value):
                    cell.fill = styles['green_fill']
                elif "Fail" in str(value):
                    cell.fill = styles['red_fill']
                elif str(value).strip().upper() in ["NR", "N/A", "Missing"]:
                    cell.fill = styles['yellow_fill']
            row += 1
            continue

        # CNUM validation
        if rule[1] == "from_value":
            if cnum_raw:
                if is_iccid_key(key):
                    norm_ml = normalize_iccid(ml_val)
                    norm = re.sub(r'\D', '', cnum_raw)
                elif is_ascii_imsi_key(key):
                    norm_ml, norm = normalize_ascii_imsi(ml_val), cnum_raw
                elif is_imsi_key(key):
                    norm_ml, norm = normalize_imsi(ml_val), re.sub(r'\D', '', cnum_raw)
                else:
                    norm_ml, norm = ml_val, cnum_raw
                if norm_ml == norm and len(norm_ml) == len(norm):
                    status["CNUM"] = "✅ Pass"
                else:
                    status["CNUM"] = "❌ Fail"
                    error_msg = f"[{key}] CNUM mismatch: ML={norm_ml}, CNUM={norm}"
                    mismatch_logs.append(error_msg)
                    validation_errors.append(error_msg)
                    overall_valid = False
            else:
                status["CNUM"] = "❌ Fail"
                error_msg = f"[{key}] CNUM Missing"
                mismatch_logs.append(error_msg)
                validation_errors.append(error_msg)
                overall_valid = False

        # SCM validation
        if rule[2] == "from_value":
            if scm_raw:
                if is_iccid_key(key):
                    norm_ml = normalize_iccid(ml_val)
                    norm = re.sub(r'\D', '', scm_raw)
                elif is_ascii_imsi_key(key):
                    norm_ml, norm = normalize_ascii_imsi(ml_val), scm_raw
                elif is_imsi_key(key):
                    norm_ml, norm = normalize_imsi(ml_val), re.sub(r'\D', '', scm_raw)
                else:
                    norm_ml, norm = ml_val, scm_raw
                if norm_ml == norm and len(norm_ml) == len(norm):
                    status["SCM"] = "✅ Pass"
                else:
                    status["SCM"] = "❌ Fail"
                    error_msg = f"[{key}] SCM mismatch: ML={norm_ml}, SCM={norm}"
                    mismatch_logs.append(error_msg)
                    validation_errors.append(error_msg)
                    overall_valid = False
            else:
                status["SCM"] = "❌ Fail"
                error_msg = f"[{key}] SCM Missing"
                mismatch_logs.append(error_msg)
                validation_errors.append(error_msg)
                overall_valid = False

        # SIM_ODA validation
        if rule[3] == "from_value":
            if sim_oda_raw:
                if is_iccid_key(key):
                    norm_ml = normalize_iccid(ml_val)
                    norm = re.sub(r'\D', '', sim_oda_raw)
                elif is_ascii_imsi_key(key):
                    norm_ml, norm = normalize_ascii_imsi(ml_val), sim_oda_raw
                elif is_imsi_key(key):
                    norm_ml, norm = normalize_imsi(ml_val), re.sub(r'\D', '', sim_oda_raw)
                else:
                    norm_ml, norm = ml_val, sim_oda_raw
                if norm_ml == norm and len(norm_ml) == len(norm):
                    status["SIM_ODA"] = "✅ Pass"
                else:
                    status["SIM_ODA"] = "❌ Fail"
                    error_msg = f"[{key}] SIM_ODA mismatch: ML={norm_ml}, SIM_ODA={norm}"
                    mismatch_logs.append(error_msg)
                    validation_errors.append(error_msg)
                    overall_valid = False
            else:
                status["SIM_ODA"] = "❌ Fail"
                error_msg = f"[{key}] SIM_ODA Missing"
                mismatch_logs.append(error_msg)
                validation_errors.append(error_msg)
                overall_valid = False

        validation_status = "✅ Pass" if overall_valid else "❌ Fail"

        # Write to Excel
        data = [
            key, ml_val, pcom_disp, cnum_disp, scm_disp, sim_oda_disp,
            status["PCOM"], status["CNUM"], status["SCM"], status["SIM_ODA"], validation_status
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = styles['thick_border']
            cell.number_format = '@'
            if "Pass" in str(value):
                cell.fill = styles['green_fill']
            elif "Fail" in str(value):
                cell.fill = styles['red_fill']
            elif str(value).strip().upper() in ["NR", "N/A", "Missing"]:
                cell.fill = styles['yellow_fill']

        row += 1

    # --- Begin processing ---
    # Initialize JIO SCM Reader
    scm_reader = None
    if scm_path and os.path.exists(scm_path):
        print(f"📊 Initializing JIO SCM Reader from {scm_path}")
        scm_reader = SCMReader(scm_path)

    image_label_map = {
        "INNER LABEL 100": image_paths[0] if len(image_paths) > 0 else None,
        "INNER LABEL 500": image_paths[1] if len(image_paths) > 1 else None,
        "OUTER LABEL 5000": image_paths[2] if len(image_paths) > 2 else None,
        "ARTWORK FRONT": image_paths[3] if len(image_paths) > 3 else None,
        "ARTWORK BACK": image_paths[4] if len(image_paths) > 4 else None,
    }

    print("Image Label Map:", image_label_map)

    # === Smart Mapping Preparation ===
    # Build a lookup cache of validated values to map raw barcodes
    validated_cache = {}
    
    # Map headings to common label fields
    mapping_hints = {
        "ICCID_CARD (2FE2)": "ICCID",
        "GLOBAL_IMSI (3031)": "IMSI",
        "HOME_IMSI (6F07)": "IMSI",
        "GLOBAL_ACC (3037)": "ACC",
        "HOME_ACC (6F78)": "ACC",
    }
    
    for k, v in results.items():
        if v and v not in ["N/A", "NR", "Missing"]:
            validated_cache[k.upper().replace(" ", "")] = v
            if k in mapping_hints:
                validated_cache[mapping_hints[k]] = v
                
    for k, files in file_values.items():
        for src, v in files.items():
            if v and v not in ["N/A", "NR", "Missing"]:
                val_clean = str(v).strip()
                validated_cache[f"{k.upper().replace(' ', '')}_{src}"] = val_clean
                
                # Broad hints
                if "ICCID" in k.upper(): validated_cache["ICCID"] = val_clean
                if "PO" in k.upper(): validated_cache["PO"] = val_clean
                if "BATCH" in k.upper(): validated_cache["BATCHNO"] = val_clean
                if "QTY" in k.upper(): validated_cache["QTY"] = val_clean
                if "EAN" in k.upper(): validated_cache["EAN"] = val_clean
                if "CIRCLE" in k.upper(): validated_cache["CIRCLE"] = val_clean
                if "PRODUCT" in k.upper(): validated_cache["PRODUCT"] = val_clean

    label_qr_data = {}
    
    # 1. Scan Images (Unified Logic for MOB, WBIOT, NBIOT)
    # We prioritize WBIOT logic as it handles XML parsing best
    for label, img_path in image_label_map.items():
        if img_path and os.path.isfile(img_path):
            # print(f"--- Scanning {label} from {img_path}")
            # Use data as-is (do not normalize keys yet, we need original formatting for headers)
            raw_qr_data = process_qr_code_wbiot(img_path) or {}
            # print(f"QR parsed data for {label}: {list(raw_qr_data.keys())}")
            label_qr_data[label] = raw_qr_data
        else:
            print(f"--- Skipping {label}, file missing or invalid: {img_path}")
            label_qr_data[label] = {}

    # Merge ARTWORK FRONT and BACK into "ARTWORK"
    artwork_merged = {}
    for key in ["ARTWORK FRONT", "ARTWORK BACK"]:
        for k, v in label_qr_data.get(key, {}).items():
            if k == "RAW_BARCODES":
                artwork_merged.setdefault("RAW_BARCODES", []).extend(v)
            elif k == "BARCODE_DATA":
                continue # Rebuild later
            else:
                artwork_merged[k] = v
    if artwork_merged:
        # Deduplicate RAW_BARCODES
        if "RAW_BARCODES" in artwork_merged:
            artwork_merged["RAW_BARCODES"] = list(set(artwork_merged["RAW_BARCODES"]))
        label_qr_data["ARTWORK"] = artwork_merged

    # 2. Build Dynamic Sections
    # Define which sections we want to report on
    target_sections = ["INNER LABEL 100", "INNER LABEL 500", "OUTER LABEL 5000", "ARTWORK"]
    label_sections = []

    def resolve_field_value(field, section_data, raw_barcodes, iccid_barcodes, section_title, used_barcodes):
        normalized_field = field.upper().replace(" ", "").replace(".", "").replace("_", "")
        value = ""

        # REMOVED: "SMART ICCID MAPPING" - We now use values directly from barcode as requested
        # The barcode field names (ICCID Start, ICCID End) are used as-is without sorting/reassignment

        # --- PRIORITY 1: STRUCTURED XML TAGS (from BARCODE_TAGS) ---
        barcode_tags = section_data.get("BARCODE_TAGS", [])
        for entry in barcode_tags:
            for k, v in entry.items():
                if k.upper().replace("_", "").replace(" ", "") == normalized_field:
                    return str(v).strip()

        # --- PRIORITY 3: FLAT XML/KV DATA (Fallback) ---
        for k, v in section_data.items():
            if k == "BARCODE_TAGS" or k == "RAW_BARCODES": continue
            if k.upper().replace("_", "").replace(" ", "") == normalized_field:
                return str(v).strip()
            
        # --- ROBUSTNESS: RAW XML REGEX EXTRACTION ---
        if raw_barcodes:
            import re
            search_tag = field.replace(" ", "[:_ ]?") 
            # Standard pattern for proper XML: <TAG>value</TAG>
            pattern = r"<(" + search_tag + r")>(.*?)</\1>"
            
            for bc in raw_barcodes:
                match = re.search(pattern, bc, re.IGNORECASE)
                if match:
                    return match.group(2).strip()
            
            # FALLBACK: Handle malformed XML like <value</TAG> (missing opening tag)
            # This matches cases where content appears before closing tag like: <MSC>MC01</MSC><URT00001234871A001</MSN1>
            for bc in raw_barcodes:
                # Look for pattern: any content followed by </MSN1> (or other TAG)
                # The malformed format is: </TAG><value_without_tag_closing
                match = re.search(r'(.*?)(</' + search_tag + r'>)', bc, re.IGNORECASE)
                if match:
                    content = match.group(1)
                    # Find content after the last > which is the value before the closing tag
                    if '>' in content:
                        value = content.split('>')[-1].strip()
                        if value and not value.startswith('<'):
                            return value
                
                # NEW: Handle case where value comes AFTER the closing tag like: </MSN1><URT00001234871A001
                # This is the pattern: </TAG><value_without_any_tag
                next_match = re.search(r'</' + search_tag + r'>\s*([^<]+)', bc, re.IGNORECASE)
                if next_match:
                    value = next_match.group(1).strip()
                    if value and not value.startswith('<'):
                        return value
                
                # NEW: Handle case where value comes BEFORE closing tag (missing opening tag): <value</MSN1>
                # This is the pattern: <value_without_tag<TAG> or <value</TAG>
                before_match = re.search(r'<([^>]+)</' + search_tag + r'>', bc, re.IGNORECASE)
                if before_match:
                    value = before_match.group(1).strip()
                    # Make sure it's not another tag
                    if value and not value.startswith('<') and '>' not in value:
                        return value

        # --- SMART MAPPING (Other Fields) ---
        
        # (ICCID block was moved to Priority 1)
        
        # 2. Match from raw barcodes using heuristics
        if raw_barcodes:
            if normalized_field == "SCANNEDDATA":
                return ", ".join(raw_barcodes)

            for bc in raw_barcodes:
                if bc in used_barcodes: continue 

                is_xml = "<" in bc and ">" in bc
                
                if normalized_field in ["PO", "PONO"] and not is_xml and len(bc) >= 8 and len(bc) <= 12 and bc.isdigit():
                    if bc.startswith("45") or not any(x.startswith("45") and x.isdigit() for x in raw_barcodes if x not in used_barcodes):
                        used_barcodes.add(bc)
                        return bc
                elif normalized_field == "EAN" and not is_xml and bc.isdigit() and len(bc) in [9, 12, 13]:
                    used_barcodes.add(bc)
                    return bc
                elif "MSN" in normalized_field and bc.startswith("URT"):
                    # Extract the counter part (A001, A002, etc) from the MSN value
                    # MSN format: URT + SKU(8) + PO(3) + Counter(A001) = 18 chars
                    if len(bc) >= 14:
                        counter_part = bc[-4:]  # Get last 4 chars like "A001"
                        if counter_part.startswith("A") and counter_part[1:].isdigit():
                            msn_counter_num = int(counter_part[1:])  # Extract 1 from "A001"
                            # Check if this matches the requested MSN number
                            if any(char.isdigit() for char in normalized_field):
                                field_num = int("".join([c for c in normalized_field if c.isdigit()]))
                                if msn_counter_num == field_num:
                                    used_barcodes.add(bc)
                                    return bc
                            else:
                                # No number in field, return first MSN found
                                used_barcodes.add(bc)
                                return bc
                    # Fallback: original logic for shorter barcodes
                    if any(char.isdigit() for char in normalized_field):
                        field_num = "".join([c for c in normalized_field if c.isdigit()])
                        if bc.endswith(field_num) or bc.endswith(field_num.zfill(2)):
                            used_barcodes.add(bc)
                            return bc
                    else:
                        used_barcodes.add(bc)
                        return bc
                elif normalized_field == "MSC" and (bc.startswith("URT") or "MC" in bc):
                    used_barcodes.add(bc)
                    return bc
                elif normalized_field == "PID":
                    if (bc.startswith("PRT") or bc.startswith("PID") or bc.startswith("JIO") or 
                        (len(bc) >= 8 and any(c.isalpha() for c in bc) and not bc.startswith("8991") and not bc.startswith("URT"))):
                        used_barcodes.add(bc)
                        return bc
                elif normalized_field == "CIRCLE" and not is_xml and len(bc) == 2 and bc.isalpha() and bc.isupper():
                    used_barcodes.add(bc)
                    return bc
                elif normalized_field == "QTY" and not is_xml and len(bc) in [2, 3, 4, 5] and bc.isdigit() and int(bc) <= 10000:
                    used_barcodes.add(bc)
                    return bc
                
        # Special case for Artwork aggregation
        if section_title == "ARTWORK" and (normalized_field == "ICCIDSTART" or normalized_field == "ICCID"):
            return ", ".join(raw_barcodes)

        return ""
    
    # ... (Skip map logic remains)
    skip_map = {
        "MOB": ["INNER_LABEL_100"], # Re-added per user request to skip 100 for MOB
        "WBIOT": ["GLOBAL_IMSI (3031)", "GLOBAL_ACC (3037)", "HOME_ACC (3037)"],
        "NBIOT": ["GLOBAL_IMSI (3031)", "GLOBAL_ACC (3037)", "HOME_ACC (3037)", 
                  "ASCII_IMSI (6F02)", "ASCII_IMSI (6F04)"]
    }

    skip_fields = set(skip_map.get(profile_type, []))

    # Field definitions
    # ... (rest of key_to_header)
    
    # ... (skipping to target sections logic)
    
    # 2. Build Dynamic Sections
    # Define which sections we want to report on
    if profile_type == "MOB":
        target_sections = ["INNER LABEL 500", "OUTER LABEL 5000", "ARTWORK"]
    else:
        target_sections = ["INNER LABEL 100", "INNER LABEL 500", "OUTER LABEL 5000", "ARTWORK"]
        
    label_sections = []

    # Priority for sorting columns (only to order them nicely if they exist)
    PRIORITY = ["ICCID Start", "ICCID End", "ICCID", "PO", "QTY", "EAN", "CIRCLE", "MSC", "MSN", "PID"]

    def get_sort_key(field):
        f_upper = field.upper().replace(" ", "")
        # Exact priority match
        for i, p in enumerate(PRIORITY):
            if f_upper == p.upper().replace(" ", ""): return i
        # MSN/MSC variants
        if "MSN" in f_upper:
            # Sort MSN1...MSN10 numerically
            nums = re.findall(r'\d+', f_upper)
            if nums: return 20 + int(nums[0])
            return 20
        if "MSC" in f_upper: return 30
        return 100 # Others at end

    for section in target_sections:
        raw_data = label_qr_data.get(section, {})
        keys = [k for k in raw_data.keys() if k not in ["RAW_BARCODES", "BARCODE_DATA"]]
        raw_barcodes = raw_data.get("RAW_BARCODES", [])
        
        # DEBUG: Check what fields we need to extract for each section
        if section == "OUTER LABEL 5000":
            print(f"[DEBUG] {section} keys (from raw_data): {keys}")
        
        # DEBUG: Print fields detected for each section
        if section == "OUTER LABEL 5000":
            print(f"[DEBUG] OUTER LABEL 5000 keys: {keys}")
            print(f"[DEBUG] OUTER LABEL 5000 raw_barcodes count: {len(raw_barcodes)}")
            print(f"[DEBUG] OUTER LABEL 5000 raw_data keys: {list(raw_data.keys())}")
            if raw_barcodes:
                print(f"[DEBUG] All raw_barcodes for OUTER LABEL 5000:")
                for i, bc in enumerate(raw_barcodes):
                    print(f"  [{i}] {bc[:150]}")
        
        # User Request: If no label data (barcodes) and no circle input, skip section
        if not keys and not raw_barcodes and not (circle_value and str(circle_value).strip()):
            print(f"[INFO] Skipping optional section {section} (empty)")
            continue
        
        display_fields = set()
        seen_normalized = set()
        
        # Pure Dynamic Mapping: Use keys exactly as found (just prettier)
        for k in keys:
            # Check for duplicates (case-insensitive)
            norm = k.upper().replace("_", "").replace(" ", "")
            if norm in seen_normalized or norm == "BARCODETAGS":
                continue
            
            # Replace underscores with spaces (e.g. ICCID_START -> ICCID START)
            disp = k.replace("_", " ")
            display_fields.add(disp)
            seen_normalized.add(norm)

        # RECOVERY: If keys are empty (parsing failed) OR even if they exist, 
        # let's scan RAW_BARCODES for tags to ensure we didn't miss anything.
        # This fixes "xml data is present but report is empty".
        if "RAW_BARCODES" in raw_data:
            for bc in raw_data["RAW_BARCODES"]:
                 # Regex to find <Tags>
                 found_tags = re.findall(r'<(\w[\w\s]*)>', bc)
                 if found_tags:
                     for tag in found_tags:
                         # Clean tag
                         clean_tag = tag.strip().replace("_", " ")
                         norm = clean_tag.upper().replace(" ", "")
                         
                         if norm in seen_normalized:
                             continue
                             
                         display_fields.add(clean_tag)
                         seen_normalized.add(norm)
                 else:
                     # CSV Fallback for Inner Label 100 (e.g. "274453, 899..., 450..., 100, KO, ...")
                     # Heuristic: If it has > 5 commas and looks like the known format
                     parts = [p.strip() for p in bc.split(',')]
                     if len(parts) >= 8 and "INNER LABEL 100" in section.upper():
                         # print(f"DEBUG: Detected CSV format for {section}: {parts}")
                         # Mapping based on user sample:
                         # 0:ID?, 1:ICCID, 2:PO, 3:QTY, 4:Circle, 5:MSN, 6:ICCID_End?, 7:EAN
                         csv_map = {
                             "ICCID Start": parts[1],
                             "PO": parts[2],
                             "QTY": parts[3],
                             "Circle": parts[4],
                             "MSN": parts[5],
                             "ICCID End": parts[6],
                             "EAN": parts[7]
                         }
                         # Add these fields to display_fields and START_DATA for resolution
                         for k, v in csv_map.items():
                             display_fields.add(k)
                             if k not in raw_data: 
                                 raw_data[k] = v
                                 raw_data[k.upper().replace(" ", "")] = v
            
            # DISPERSED BARCODE ASSEMBLY (Fallback for Inner 100 if no CSV/XML)
            # If we still have no significant fields and it is Inner 100
            if "INNER LABEL 100" in section.upper() and not any("ICCID" in f.upper() for f in display_fields):
                 # print(f"DEBUG: Attempting Dispersed Component Assembly for {section}")
                 dispersed_map = {}
                 iccid_candidates = []

                 for bc in raw_data["RAW_BARCODES"]:
                     bc = bc.strip()
                     # Heuristics
                     if bc.startswith("89") and len(bc) >= 19 and len(bc) <= 20 and bc.isdigit():
                         iccid_candidates.append(bc)
                     elif bc.startswith("450") and len(bc) >= 9 and bc.isdigit():
                         dispersed_map["PO"] = bc
                     elif bc.startswith("499") and len(bc) >= 9 and bc.isdigit():
                         dispersed_map["EAN"] = bc
                     elif bc == "100":
                         dispersed_map["QTY"] = bc
                     elif bc.startswith("URT"):
                         dispersed_map["MSN"] = bc
                     elif len(bc) == 2 and bc.isalpha():
                         dispersed_map["CIRCLE"] = bc
                 
                 # Logic to determine Start/End from candidates
                 if iccid_candidates:
                     # No sorting - pass as discovered, validator will handle alignment
                     dispersed_map["ICCID Start"] = iccid_candidates[0]
                     if len(iccid_candidates) > 1:
                         dispersed_map["ICCID End"] = iccid_candidates[1] # Take second one if exists

                 # If we found at least ICCID, populate
                 if "ICCID Start" in dispersed_map:
                     for k, v in dispersed_map.items():
                         display_fields.add(k)
                         if k not in raw_data:
                             raw_data[k] = v
                             raw_data[k.upper().replace(" ", "")] = v

        # FORCE RENAME: For Artwork sections, if we have "Scanned Data" or it's empty, use "ICCID Start"
        # This overrides previous default/fallback logic to satisfy user request.
        if "ARTWORK" in section.upper():
            if not display_fields or "Scanned Data" in display_fields:
                if "Scanned Data" in display_fields:
                    display_fields.remove("Scanned Data")
                display_fields.add("ICCID Start")

        if display_fields:
            sorted_fields = sorted(list(display_fields), key=get_sort_key)
            label_sections.append((section, sorted_fields))
        else:
            # No fields found. User requested ONLY what is found.
            # Fallback to "Scanned Data" if only raw barcodes exist.
            if "RAW_BARCODES" in raw_data and raw_data["RAW_BARCODES"]:
                 # Just show a generic field so we can dump the raw value
                 # User Request: Use "ICCID Start" for Artwork
                 if "ARTWORK" in section:
                     label_sections.append((section, ["ICCID Start"]))
                 else:
                     label_sections.append((section, ["Scanned Data"]))
            else:
                 # Truly empty
                 # For MOB if 100 label is scanned but we want to ignore it?
                 if section == "INNER LABEL 100" and profile_type == "MOB":
                     continue # Should have been filtered by target_sections, but safe guard
                 label_sections.append((section, []))

    # Styles
    label_fill = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
    label_font = Font(color='FFFFFF', bold=True, size=12)
    label_align = Alignment(horizontal='left', vertical='center')
    bold_border = Border(
        left=Side(style='thick'), right=Side(style='thick'),
        top=Side(style='thick'), bottom=Side(style='thick')
    )

    # Start row (after main table)
    label_start_row = row + 1  # 'row' must already be defined

    # Write sections
    current_section_title = None
    current_section_data = {}
    section_validation = {"status": "PASS", "field_status": {}}
    
    # USER REQUEST: If OUTER LABEL 5000 fails (e.g. Count Mismatch), propagate to other labels
    global_outer_count_fail = False
    
    # ============================================================
    # USER REQUEST: Extract reference values from Inner Label 500
    # These will be used to validate Inner Label 100
    # ============================================================
    inner_label_500_reference = {
        "iccid_start": None,
        "sku_code": None,
        "ean": None,
        "po": None
    }
    
    # Get Inner Label 500 data for cross-reference
    if "INNER LABEL 500" in label_qr_data:
        inner_500_data = label_qr_data.get("INNER LABEL 500", {})
        inner_500_raw_barcodes = inner_500_data.get("RAW_BARCODES", [])
        
        # Extract ICCID Start from Inner Label 500
        for f in ["ICCID Start", "ICCIDSTART", "ICCID_START"]:
            if f in inner_500_data and inner_500_data[f]:
                inner_label_500_reference["iccid_start"] = str(inner_500_data[f]).strip()
                break
            # Try from raw barcodes
            for bc in inner_500_raw_barcodes:
                if bc.startswith("8991") and len(bc) >= 18:
                    inner_label_500_reference["iccid_start"] = bc.strip()
                    break
        
        # Extract SKU Code/EAN from Inner Label 500
        for f in ["SKU Code", "SKUCODE", "EAN"]:
            if f in inner_500_data and inner_500_data[f]:
                val = str(inner_500_data[f]).strip()
                if "SKU" in f.upper():
                    inner_label_500_reference["sku_code"] = val
                elif "EAN" in f.upper():
                    inner_label_500_reference["ean"] = val
        
        # Extract PO from Inner Label 500
        for f in ["PO", "PONO"]:
            if f in inner_500_data and inner_500_data[f]:
                inner_label_500_reference["po"] = str(inner_500_data[f]).strip()
                break
        
        print(f"[INFO] Inner Label 500 Reference - ICCID Start: {inner_label_500_reference['iccid_start']}, SKU: {inner_label_500_reference['sku_code']}, EAN: {inner_label_500_reference['ean']}, PO: {inner_label_500_reference['po']}")
    
    if scm_reader and "OUTER LABEL 5000" in label_qr_data:
        # print("[GLOBAL] Pre-validating Outer Label for failure propagation...")
        outer_raw_data = label_qr_data.get("OUTER LABEL 5000", {})
        outer_raw_barcodes = outer_raw_data.get("RAW_BARCODES", [])
        outer_iccid_barcodes = [bc for bc in outer_raw_barcodes if bc.startswith("8991") and len(bc) >= 18]
        
        outer_fields = []
        for section_title, fields in label_sections:
            if section_title == "OUTER LABEL 5000":
                outer_fields = fields
                break
        
        if outer_fields:
            temp_outer_data = {}
            for f in outer_fields:
                v = resolve_field_value(f, outer_raw_data, outer_raw_barcodes, outer_iccid_barcodes, "OUTER LABEL 5000", set())
                # Simplified normalization for pre-pass
                norm_f = f.upper().replace(" ", "").replace("_", "")
                validator_key = f
                if "ICCIDSTART" in norm_f: validator_key = "ICCID Start"
                elif "ICCIDEND" in norm_f: validator_key = "ICCID End"
                elif "QTY" in norm_f: validator_key = "QTY"
                temp_outer_data[validator_key] = v
            
            ov = validate_outer_label_5000(temp_outer_data, scm_reader, gui_circle=circle_value)
            # USER REQUEST: Only propagate if it's a COUNT/CONTINUITY issue
            # Check details for specific keywords to avoid triggering on simple PO/EAN mismatches
            ov_details = ov.get("details", [])
            is_count_issue = any("Count Mismatch" in d or "Continuity Error" in d for d in ov_details)
            
            if ov.get("status") == "FAIL":
                if is_count_issue or ov.get("field_status", {}).get("QTY") == "FAIL" or ov.get("field_status", {}).get("ICCID End") == "FAIL":
                    global_outer_count_fail = True
                    print("[GLOBAL] Outer Label count failure detected. Will propagate to all sections.")

    for idx, (section_title, fields) in enumerate(label_sections):
        # Section header
        for col in range(1, 4):
            cell = ws.cell(row=label_start_row, column=col)
            cell.fill = label_fill
            cell.font = label_font
            cell.alignment = label_align
            cell.border = bold_border

        ws.cell(row=label_start_row, column=1, value=section_title)
        ws.cell(row=label_start_row, column=2, value="Values")
        ws.cell(row=label_start_row, column=3, value="Status")

        # Get normalized data
        section_data_raw = label_qr_data.get(section_title, {})
        raw_barcodes = section_data_raw.get("RAW_BARCODES", [])
        
        # Track used barcodes in this section to avoid mapping the same barcode to multiple fields
        used_barcodes = set()
        
        # Identify all unique ICCID-like barcodes in this image and sort them numerically
        # Deduplication is necessary because multiple scanner filters may pick up the same barcode.
        # Sorting ensures Start ICCID (lower) and End ICCID (higher) are mapped consistently.
        # IMPROVED: Scan INSIDE all barcodes (including XML) for 8991 strings
        raw_iccids = []
        for bc in raw_barcodes:
            if bc.startswith("8991") and len(bc) >= 18:
                raw_iccids.append(re.sub(r'\D', '', bc)[:20]) # Ensure clean digits
            else:
                # Search for 8991 strings inside XML or complex barcodes
                matches = re.findall(r'8991\d{14,16}', bc)
                raw_iccids.extend(matches)
        
        iccid_barcodes = sorted(list(set(raw_iccids)))
        for bc in iccid_barcodes: used_barcodes.add(bc)
        
        section_data = {
            k.upper().replace(" ", "").replace(".", ""): v
            for k, v in section_data_raw.items()
            if k not in ["RAW_BARCODES", "BARCODE_DATA"]
        }

        # JIO VALIDATION PRE-PASS: Validate entire section before writing fields
        # This ensures corrected alignment (swaps) are reflected in the report values.
        section_validation = {"status": "PASS", "field_status": {}, "details": []}
        if scm_reader and section_title in ["INNER LABEL 100", "INNER LABEL 500", "OUTER LABEL 5000", "ARTWORK FRONT", "ARTWORK BACK", "ARTWORK"]:
            current_section_data = {}
            
            # First, populate current_section_data with resolved field values
            for f in fields:
                v = resolve_field_value(f, section_data_raw, raw_barcodes, iccid_barcodes, section_title, set())
                # Normalize keys for validator
                norm_f = f.upper().replace(" ", "").replace("_", "")
                validator_key = f
                if "ICCIDSTART" in norm_f: validator_key = "ICCID Start"
                elif "ICCIDEND" in norm_f: validator_key = "ICCID End"
                elif "QTY" in norm_f: validator_key = "QTY"
                elif "PO" in norm_f: validator_key = "PO"
                elif "EAN" in norm_f: validator_key = "EAN"
                elif "MSC" in norm_f: 
                    # Preserve suffix if it's like MSC1, MSC2 etc
                    validator_key = norm_f if any(c.isdigit() for c in norm_f) else "MSC"
                elif "MSN" in norm_f: 
                    # Preserve suffix if it's like MSN1, MSN2 etc
                    validator_key = norm_f if any(c.isdigit() for c in norm_f) else "MSN"
                elif "CIRCLE" in norm_f: validator_key = "Circle"
                elif "PID" in norm_f:
                    # Preserve suffix if it's like PID1, PID2 etc
                    validator_key = norm_f if any(c.isdigit() for c in norm_f) else "PID"
                current_section_data[validator_key] = v
            
            # ============================================================
            # USER REQUEST: Inner Label 100 Cross-Reference Logic
            # Get ICCID Start from Inner Label 500 and ALWAYS use it as ICCID Start
            # This MUST run BEFORE the debug section to avoid being overwritten
            # ============================================================
            if section_title == "INNER LABEL 100" and inner_label_500_reference.get("iccid_start"):
                reference_iccid = inner_label_500_reference["iccid_start"]
                reference_sku = inner_label_500_reference.get("sku_code")
                reference_ean = inner_label_500_reference.get("ean")
                reference_po = inner_label_500_reference.get("po")
                
                print(f"[INFO] Inner Label 100 Cross-Reference: Using ICCID {reference_iccid} from Inner Label 500 as START")
                
                # ALWAYS set the ICCID from Inner Label 500 as ICCID Start
                current_section_data["ICCID Start"] = reference_iccid
                
                # Find the OTHER ICCID in Inner Label 100 (not the matched one) for ICCID End
                all_iccids_in_100 = []
                ref_clean = re.sub(r'\D', '', str(reference_iccid))[:20]
                
                for bc in raw_barcodes:
                    bc_clean = re.sub(r'\D', '', str(bc))[:20]
                    if bc_clean.startswith("8991") and len(bc_clean) >= 18:
                        if bc_clean != ref_clean:
                            all_iccids_in_100.append(bc.strip())
                
                if all_iccids_in_100:
                    current_section_data["ICCID End"] = all_iccids_in_100[0]
                    print(f"[INFO] Inner Label 100 ICCID End set to: {all_iccids_in_100[0]}")
                else:
                    current_section_data["ICCID End"] = ""

            # USER REQUEST: Debug printing for ALL labels showing Scanned vs Logic
            if section_title in ["INNER LABEL 100", "INNER LABEL 500", "OUTER LABEL 5000"]:
                print(f"\n" + "="*50)
                print(f"--- DEBUG: {section_title} DATA ---")
                
                # CRITICAL: Store original scanned ICCID values BEFORE any validation
                # These values must be displayed exactly as scanned, without any swapping or modification
                import copy
                original_iccid_start = current_section_data.get("ICCID Start", "")
                original_iccid_end = current_section_data.get("ICCID End", "")
                
                # We validate here once to get the logic/expected values for the debug print
                # Use deep copy to prevent modification of original data
                temp_val = {}
                if section_title == "OUTER LABEL 5000":
                    temp_val = validate_outer_label_5000(copy.deepcopy(current_section_data), scm_reader, gui_circle=circle_value)
                else:
                    lqty = "100" if "100" in section_title else "500"
                    temp_val = validate_jio_label(lqty, copy.deepcopy(current_section_data), scm_reader, gui_circle=circle_value)
                
                # Restore original scanned values for display
                if original_iccid_start:
                    current_section_data["ICCID Start"] = original_iccid_start
                if original_iccid_end:
                    current_section_data["ICCID End"] = original_iccid_end
                
                for k, v in current_section_data.items():
                    # Skip empty MSN/PID/MSC sub-blocks (like MSN2-MSN10 when empty)
                    # Only show fields that have values or are explicitly validated
                    field_upper = str(k).upper()
                    is_msn_block = "MSN" in field_upper and any(c.isdigit() for c in str(k))
                    is_msc_block = "MSC" in field_upper and any(c.isdigit() for c in str(k))
                    is_pid_block = "PID" in field_upper and any(c.isdigit() for c in str(k))
                    
                    # Skip empty sub-blocks
                    if not v and (is_msn_block or is_msc_block or is_pid_block):
                        continue
                        
                    status = "✅" if temp_val.get("field_status", {}).get(k) == "PASS" else "❌"
                    print(f"   {str(k).upper():<12}: {v:<25} {status}")
                
                if temp_val.get("details"):
                    print("   LOGIC DETAILS:")
                    for d in temp_val["details"]:
                        print(f"      - {d}")
                print("="*50 + "\n")

            start_iccid = current_section_data.get("ICCID Start", "Unknown")
            # print(f"[SCAN] Validating {section_title} against SCM (Starting ICCID: {start_iccid})...")
            
            # ============================================================
            # USER REQUEST: Inner Label 100 Cross-Reference Logic
            # Get ICCID Start from Inner Label 500 and ALWAYS use it as ICCID Start in Inner Label 100
            # Inner Label 100 has only barcode values without tags, so we need to identify
            # the starting ICCID by matching with Inner Label 500
            # ============================================================
            if section_title == "INNER LABEL 100" and inner_label_500_reference.get("iccid_start"):
                reference_iccid = inner_label_500_reference["iccid_start"]
                reference_sku = inner_label_500_reference.get("sku_code")
                reference_ean = inner_label_500_reference.get("ean")
                reference_po = inner_label_500_reference.get("po")
                
                print(f"[INFO] Inner Label 100 Cross-Reference: Using ICCID {reference_iccid} from Inner Label 500 as START")
                
                # ALWAYS set the ICCID from Inner Label 500 as ICCID Start
                # This is the authoritative reference value
                current_section_data["ICCID Start"] = reference_iccid
                
                # Find the OTHER ICCID in Inner Label 100 (not the matched one) for ICCID End
                # Search through raw barcodes to find ICCIDs
                all_iccids_in_100 = []
                ref_clean = re.sub(r'\D', '', str(reference_iccid))[:20]
                
                for bc in raw_barcodes:
                    bc_clean = re.sub(r'\D', '', str(bc))[:20]
                    # Check if this is a valid ICCID (starts with 8991 and ~20 digits)
                    if bc_clean.startswith("8991") and len(bc_clean) >= 18:
                        # Skip the reference ICCID, add the others
                        if bc_clean != ref_clean:
                            all_iccids_in_100.append(bc.strip())
                
                # If we found another ICCID, use it as ICCID End
                if all_iccids_in_100:
                    # Take the first other ICCID found (should be only 1 for Inner Label 100)
                    current_section_data["ICCID End"] = all_iccids_in_100[0]
                    print(f"[INFO] Inner Label 100 ICCID End set to: {all_iccids_in_100[0]}")
                else:
                    # Clear the ICCID End if no other ICCID found
                    current_section_data["ICCID End"] = ""
                
                # Cross-check SKU Code and EAN with Inner Label 500
                # If Inner Label 100 has different SKU/EAN, flag as issue
                current_sku = current_section_data.get("EAN") or current_section_data.get("SKU Code") or current_section_data.get("SKUCODE")
                current_ean = current_section_data.get("EAN")
                
                # Check if SKU/EAN matches reference
                if reference_sku and current_sku and reference_sku != current_sku:
                    print(f"[WARN] Inner Label 100 SKU Mismatch: Label={current_sku}, Reference={reference_sku}")
                    # Add to validation details (will be handled below)
                
                if reference_ean and current_ean and reference_ean != current_ean:
                    print(f"[WARN] Inner Label 100 EAN Mismatch: Label={current_ean}, Reference={reference_ean}")
                    # Add to validation details (will be handled below)
            
            # CRITICAL: Store original scanned ICCID values BEFORE validation
            # These values must be displayed exactly as scanned, without any swapping or modification
            import copy
            original_iccid_start = current_section_data.get("ICCID Start", "")
            original_iccid_end = current_section_data.get("ICCID End", "")
            
            if section_title == "OUTER LABEL 5000":
                section_validation = validate_outer_label_5000(copy.deepcopy(current_section_data), scm_reader, gui_circle=circle_value)
            else:
                label_qty = "100" if "100" in section_title else "500" if "500" in section_title else "1"
                section_validation = validate_jio_label(label_qty, copy.deepcopy(current_section_data), scm_reader, gui_circle=circle_value)
            
            # ============================================================
            # USER REQUEST: Add cross-reference validation results for Inner Label 100
            # Check if SKU/EAN matches Inner Label 500 reference
            # ============================================================
            if section_title == "INNER LABEL 100" and inner_label_500_reference.get("iccid_start"):
                reference_sku = inner_label_500_reference.get("sku_code")
                reference_ean = inner_label_500_reference.get("ean")
                reference_po = inner_label_500_reference.get("po")
                
                # Check EAN match
                current_ean = current_section_data.get("EAN")
                if reference_ean and current_ean and reference_ean != current_ean:
                    section_validation["field_status"]["EAN"] = "FAIL"
                    section_validation["status"] = "FAIL"
                    section_validation["details"].append(f"EAN Mismatch with Inner Label 500: Label={current_ean}, Inner500={reference_ean}")
                
                # Check SKU Code (could be in EAN field or separate)
                current_sku = current_section_data.get("EAN")  # Sometimes SKU is in EAN field
                if reference_sku and current_sku and reference_sku != current_sku:
                    # Also check if EAN matches
                    if not (current_ean and reference_ean and current_ean == reference_ean):
                        section_validation["field_status"]["SKU Code"] = "FAIL"
                        section_validation["status"] = "FAIL"
                        section_validation["details"].append(f"SKU Code Mismatch with Inner Label 500: Label={current_sku}, Inner500={reference_sku}")
                
                # Check PO match
                current_po = current_section_data.get("PO")
                if reference_po and current_po and reference_po != current_po:
                    section_validation["field_status"]["PO"] = "FAIL"
                    section_validation["status"] = "FAIL"
                    section_validation["details"].append(f"PO Mismatch with Inner Label 500: Label={current_po}, Inner500={reference_po}")
            
            # Restore original scanned values for report
            if original_iccid_start:
                current_section_data["ICCID Start"] = original_iccid_start
            if original_iccid_end:
                current_section_data["ICCID End"] = original_iccid_end
            
            # Map corrected values back to section_data_raw for resolve_field_value to use
            if section_validation.get("status") == "PASS" or True: # Always sync even on fail for consistency
                if "ICCID Start" in current_section_data:
                    section_data_raw["ICCID Start"] = current_section_data["ICCID Start"]
                if "ICCID End" in current_section_data:
                    section_data_raw["ICCID End"] = current_section_data["ICCID End"]

            if section_validation["status"] == "FAIL":
                # print(f"   [FAIL] {section_title} Validation Failed: {len(section_validation['details'])} issues found.")
                for detail in section_validation["details"]:
                    print(f"      - {detail}")
                    validation_errors.append(f"[{section_title}] {detail}")
            else:
                print(f"   [PASS] {section_title} Validation Passed.")

        # Section fields: Write to Excel
        for field in fields:
            # Resolve value (will use updated section_data_raw if alignment found)
            value = resolve_field_value(field, section_data_raw, raw_barcodes, iccid_barcodes, section_title, used_barcodes)
            
            # --- STATUS DETERMINATION ---
            field_status = None
            if scm_reader and section_title in ["INNER LABEL 100", "INNER LABEL 500", "OUTER LABEL 5000", "ARTWORK FRONT", "ARTWORK BACK", "ARTWORK"]:
                fs_dict = section_validation.get("field_status", {})
                field_status = fs_dict.get(field)
                if field_status is None:
                    norm_field = field.upper().replace(" ", "").replace("_", "")
                    for fs_key, fs_val in fs_dict.items():
                        if fs_key.upper().replace(" ", "").replace("_", "") == norm_field:
                            field_status = fs_val
                            break
                
                if field_status is None:
                    # USER REQUEST: Don't force FAIL on everything. Default to PASS if not graded.
                    field_status = "PASS"
                
                # USER REQUEST: Only propagate to MSC/PID for outer label issues
                # QTY should only fail if the specific label's batch has an issue
                # MSN blocks are validated independently
                if global_outer_count_fail and section_title == "OUTER LABEL 5000" and any(x in field.upper() for x in ["MSC", "PID"]) and "MSN" not in field.upper():
                    field_status = "FAIL"

            # User Request: If value is empty and it passed (or wasn't explicitly graded as FAIL), skip it if it's an optional sub-block
            # This specifically removes MSN2-MSN10 from Golden Sample reports.
            if not str(value).strip() and field_status == "PASS" and "ARTWORK" not in section_title:
                # If it's a numbered sub-block (MSN1-10, PID1-10, MSC1-10 etc), skip it if empty
                if any(x in field.upper() for x in ["MSN", "MSC", "PID"]) and any(c.isdigit() for c in field):
                   continue

            label_start_row += 1
            ws.cell(row=label_start_row, column=1, value=field).border = bold_border
            
            cell = ws.cell(row=label_start_row, column=2, value=value)
            cell.border = bold_border
            cell.number_format = '@' 

            # --- STATUS MAPPING ---
            status_cell = ws.cell(row=label_start_row, column=3)
            status_cell.border = bold_border
            
            if field_status is not None:
                if field_status == "PASS":
                    status_cell.value = "✅ Pass"
                    status_cell.fill = styles['green_fill']
                else:
                    # USER REQUEST: Make it clear WHY it failed with meaningful information
                    reason = ""
                    field_details = section_validation.get("details", [])
                    
                    # Priority 1: Look for exact field name match
                    field_upper = field.upper().replace(" ", "")
                    for d in field_details:
                        d_upper = d.upper().replace(" ", "")
                        # Check if this detail is about this specific field
                        if field_upper in d_upper or (field == "QTY" and ("CONTINUITY" in d_upper or "COUNT" in d_upper)):
                            # Extract the meaningful part
                            if "not found in SCM" in d:
                                reason = f" (Not found in SCM)"
                            elif "Mismatch" in d and ":" in d:
                                # Extract just the key info after the colon
                                parts = d.split(":")
                                if len(parts) > 1:
                                    detail = parts[-1].strip()
                                    # Shorten if needed
                                    if len(detail) > 40:
                                        detail = detail[:37] + "..."
                                    reason = f" ({detail})"
                                else:
                                    reason = f" ({d[:40]}...)" if len(d) > 40 else f" ({d})"
                            elif "Batch Broken" in d or "Batch Count" in d:
                                reason = " (Batch Count Broken)"
                            else:
                                # Generic detail
                                reason = f" ({d[:40]}...)" if len(d) > 40 else f" ({d})"
                            break
                    
                    # Priority 2: Global outer batch error
                    if not reason and global_outer_count_fail and any(x in field.upper() for x in ["MSC", "MSN", "QTY", "PID"]):
                        reason = " (Outer Batch Error)"
                        
                    status_cell.value = f"❌ Fail{reason}"
                    status_cell.fill = styles['red_fill']
            else:
                status_cell.value = "N/A"
                status_cell.fill = styles['yellow_fill']


        # Add space after section (except last one)
        if idx < len(label_sections) - 1:
            label_start_row += 2

    # --- Auto-Fit Columns Based on Actual Content ---

    for col_idx, _ in enumerate(headers, 1):  # Start at 1
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, label_start_row + 1)
        )
        ws.column_dimensions[col_letter].width = max_len + 8

    # --- Save Report ---
    
    # Update Final Verification Status after all validation is complete
    update_final_verification_status(ws, validation_errors)

    # Extract ICCID for filename (from SCM or Machine Log)
    iccid = file_values.get("ICCID_CARD (2FE2)", {}).get("SCM")
    if not iccid:
        iccid = results.get("ICCID_CARD (2FE2)", "Unknown_ICCID")
    
    # Extract SOF number from folder path for report naming
    folder_path = os.path.dirname(filepath) if filepath else ""
    sof_number = extract_sof_number(folder_path)
    
    # Extract PO number for report naming (from multiple sources)
    po_number = None
    
    # Try from validated_cache
    if not po_number:
        po_number = validated_cache.get("PO")
    
    # Try from file_values
    if not po_number:
        for key, files in file_values.items():
            if "PO" in key.upper():
                for src, val in files.items():
                    if val and val not in ["N/A", "NR", "Missing"]:
                        po_number = str(val).strip()
                        break
                if po_number:
                    break
    
    # After saving report
    report_path = save_report(wb, filepath, pcom_path, iccid, sof_number)
    print(f"\n" + "="*80)
    print(f"REPORT GENERATION COMPLETE")
    print(f"Report saved at: {report_path}")
    print(f"Total validation errors: {len(validation_errors)}")
    print("="*80)
    
    return report_path, validation_errors