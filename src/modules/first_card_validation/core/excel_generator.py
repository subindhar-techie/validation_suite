import os
import re
import win32com.client as win32
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
import getpass
from datetime import datetime
import csv

# ============================================================
# HELPER FUNCTIONS FOR FINAL VERIFICATION REPORT CARD
# ============================================================

def extract_sof_number(folder_path):
    """
    Extract SOF Number from folder name.
    If folder name contains RTLP prefix, use the complete RTLP code (e.g., RTLP10090).
    Folder names follow format like "RTLP10086 NBIOT RJ 50K" or "RTLP10090 - WBIoT BR 250K"
    """
    if not folder_path or not os.path.exists(folder_path):
        return ""
    
    folder_name = os.path.basename(folder_path)
    
    # Check for RTLP prefix
    rtlp_match = re.search(r'(RTLP\d+)', folder_name, re.IGNORECASE)
    if rtlp_match:
        return rtlp_match.group(1)
    
    return ""

def extract_total_quantity(folder_path):
    """
    Extract Total Quantity from folder name.
    For example in "RTLP10086 NBIOT RJ 50K", 50K is total quantity.
    In "RTLP10090 - WBIoT BR 250K", 250K is total quantity.
    """
    if not folder_path or not os.path.exists(folder_path):
        return ""
    
    folder_name = os.path.basename(folder_path)
    
    # Look for patterns like "50K", "250K", etc.
    qty_match = re.search(r'(\d+)\s*K', folder_name, re.IGNORECASE)
    if qty_match:
        return qty_match.group(1) + "K"
    
    return ""

def extract_chip_code(sim_oda_path):
    """
    Extract chip code from SIM ODA file by extracting the chip code inside Chip("code") format.
    For example Chip("TSS380A1") yields chip code TSS380A1.
    """
    if not sim_oda_path or not os.path.exists(sim_oda_path):
        return ""
    
    try:
        with open(sim_oda_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        
        # Look for Chip("code") pattern
        chip_match = re.search(r'Chip\("([^"]+)"\)', content)
        if chip_match:
            return chip_match.group(1)
        
        # Also try without quotes: Chip(code)
        chip_match = re.search(r'Chip\(([^)]+)\)', content)
        if chip_match:
            return chip_match.group(1)
            
    except Exception as e:
        print(f"Error extracting chip code: {e}")
    
    return ""

def extract_po_and_batch_from_scm(scm_path):
    """
    Extract PO Number and Batch Number from SCM file.
    - PO Number: Extract from PONUM field
    - Batch Number: Read BATCHNO column from second line onwards (first line is header)
    - Batch Quantity: Count number of data lines
    Returns: (po_number, batch_number, batch_quantity)
    """
    po_number = ""
    batch_number = ""
    batch_quantity = 0
    
    if not scm_path or not os.path.exists(scm_path):
        return po_number, batch_number, batch_quantity
    
    try:
        with open(scm_path, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.reader(f, delimiter='\t')
            
            # Read header line
            header_line = next(reader, None)
            if not header_line:
                return po_number, batch_number, batch_quantity
            
            # Normalize headers
            headers = [h.strip().upper().replace("_", "").replace(" ", "") for h in header_line]
            
            # Find column indices
            po_col = -1
            batch_col = -1
            
            for idx, h in enumerate(headers):
                if h in ["PONUM", "PO", "PURCHASEORDER", "PONO"]:
                    po_col = idx
                elif h in ["BATCHNO", "BATCH", "BATCHNUM"]:
                    batch_col = idx
            
            # Read data lines
            for row in reader:
                if not row:
                    continue
                
                # Pad row if needed
                if len(row) < len(headers):
                    row += [""] * (len(headers) - len(row))
                
                batch_quantity += 1
                
                # Get PO from first data line
                if not po_number and po_col >= 0 and po_col < len(row):
                    po_number = row[po_col].strip()
                
                # Get Batch Number from first data line
                if not batch_number and batch_col >= 0 and batch_col < len(row):
                    batch_number = row[batch_col].strip()
                
                # If we have both, we can stop
                if po_number and batch_number:
                    break
                    
    except Exception as e:
        print(f"Error extracting PO/Batch from SCM: {e}")
    
    return po_number, batch_number, batch_quantity

def extract_header_info_from_cnum(cnum_path):
    """
    Extract header information from CNUM file.
    The header section contains:
    - PO Number
    - Batch No
    - SIM Quantity
    - Circle
    - SKU
    Returns: dict with po_number, batch_number, sim_quantity
    """
    info = {
        'po_number': '',
        'batch_number': '',
        'sim_quantity': 0
    }
    
    if not cnum_path or not os.path.exists(cnum_path):
        return info
    
    try:
        with open(cnum_path, 'r', encoding='utf-8', errors='ignore') as f:
            lines = [line.strip() for line in f.readlines()[:15]]
        
        for line in lines:
            if any(key in line for key in ["PO Number:", "PO NO:"]):
                info['po_number'] = line.split(":")[1].strip()
            elif any(key in line for key in ["Batch No:", "Batch NO:"]):
                info['batch_number'] = line.split(":")[1].strip()
            elif "SIM Quantity:" in line:
                try:
                    info['sim_quantity'] = int(line.split(":")[1].strip())
                except ValueError:
                    pass
                    
    except Exception as e:
        print(f"Error extracting header info from CNUM: {e}")
    
    return info


# ============================================================
# EXCEL REPORT FUNCTIONS
# ============================================================

# def protect_excel_file(filepath, password):
#     excel = win32.gencache.EnsureDispatch('Excel.Application')
#     excel.DisplayAlerts = False
#     wb = excel.Workbooks.Open(filepath)
#     wb.Password = password  # Set open password
#     wb.Save()
#     wb.Close()
#     excel.Quit()


def insert_image(ws, image_path, cell_location):
    try:
        print(f"Inserting image at {cell_location} from path: {image_path}")
        img = ExcelImage(image_path)  # For openpyxl image insertion
        img.width = 300  # Set image width (optional)
        img.height = 200  # Set image height (optional)
        ws.add_image(img, cell_location)  # Insert the image into the worksheet
        print(f"Image inserted at {cell_location}")
    except Exception as e:
        print(f"Error inserting image: {e}")



def save_report(wb, ml_path, pcom_path, iccid=None, sof_number=None):
    try:
        # Create final report name using ICCID if available
        if iccid:
            # Clean ICCID for filename (remove any prefixes)
            clean_iccid = str(iccid).replace("ICCID_", "").strip()
            # Include SOF number if available
            if sof_number:
                report_name = f"{clean_iccid}_{sof_number}_validation_report.xlsx"
            else:
                report_name = f"{clean_iccid}_validation_report.xlsx"
        else:
            # Fallback: Extract base filename (e.g., "Log_98195808050716183064")
            base_name = os.path.splitext(os.path.basename(ml_path))[0]

            # Remove "Log_" prefix if present
            if base_name.lower().startswith("log_"):
                raw_number = base_name[4:]
            else:
                raw_number = base_name

            # --- Pairwise swap each 2 digits ---
            swapped_number = ""
            for i in range(0, len(raw_number), 2):
                if i + 1 < len(raw_number):
                    swapped_number += raw_number[i+1] + raw_number[i]
                else:
                    swapped_number += raw_number[i]  # last digit if odd length

            # Create final report name
            # Include SOF number if available
            if sof_number:
                report_name = f"{swapped_number}_{sof_number}_validation_report.xlsx"
            else:
                report_name = f"{swapped_number}_validation_report.xlsx"

        # Save in PCOM folder if available, else Desktop
        if pcom_path and os.path.exists(os.path.dirname(pcom_path)):
            pcom_folder = os.path.dirname(pcom_path)
            report_path = os.path.join(pcom_folder, report_name)
        else:
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            report_path = os.path.join(desktop, report_name)

        wb.save(report_path)
        print(f"✅ Report saved successfully: {report_path}")
        return report_path

    except Exception as e:
        print(f"❌ Failed to save report: {e}")
        return None


def setup_excel_styles():
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    thick_border = Border(left=Side(style='thick'), right=Side(style='thick'),
                          top=Side(style='thick'), bottom=Side(style='thick'))
    dark_blue_fill = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    
    return {
        'yellow_fill': yellow_fill,
        'green_fill': green_fill,
        'red_fill': red_fill,
        'thick_border': thick_border,
        'dark_blue_fill': dark_blue_fill,
        'header_font': header_font
    }

def setup_excel_headers(ws, styles, operator_name="JIO", folder_path="", scm_path="", sim_oda_path="", circle_value="", validation_errors=None, cnum_path="", perso_script_path=None):
    """
    Setup Excel headers with Final Verification Report Card values.
    
    Parameters:
    - ws: Worksheet object
    - styles: Excel styles dictionary
    - operator_name: Fixed operator name (e.g., "JIO")
    - folder_path: Path to folder (for extracting SOF Number and Total QTY)
    - scm_path: Path to SCM file (for extracting Batch Number)
    - sim_oda_path: Path to SIM ODA file (for extracting Chip code)
    - circle_value: Circle value from user input
    - validation_errors: List of validation errors (to determine final status)
    - cnum_path: Path to CNUM file (for extracting SIM Quantity as Batch QTY and PO Number)
    - perso_script_path: Path to Perso Script file
    """
    # First Card Validation Report section
    ws["A1"] = "First Card Validation Report"
    ws["A1"].fill = styles['dark_blue_fill']
    ws["A1"].font = Font(bold=True, color='FFFFFF', size=14)
    ws["A1"].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells("A1:B1")  # Adjust range as needed

    username = getpass.getuser()
    timestamp = datetime.today().strftime('%A, %d %B %Y %I:%M %p')

    # Write date to D1
    ws["D1"] = f"Date : {timestamp}"
    ws["D1"].fill = styles['dark_blue_fill']
    ws["D1"].font = Font(bold=True, color='FFFFFF', size=14)
    ws["D1"].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells("D1:E1")  # Adjust range as needed

    # Write username to D2
    ws["D2"] = f"User : {username}"
    ws["D2"].fill = styles['dark_blue_fill']
    ws["D2"].font = Font(bold=True, color='FFFFFF', size=14)
    ws["D2"].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells("D2:E2")  # Adjust range as needed

    ws.row_dimensions[1].height = 30  # Set height in points

    # Extract values for the header section
    sof_number = extract_sof_number(folder_path) if folder_path else ""
    total_qty = extract_total_quantity(folder_path) if folder_path else ""
    chip_code = extract_chip_code(sim_oda_path) if sim_oda_path else ""
    
    # Extract PO Number and Batch Number from CNUM file header (primary source)
    cnum_info = extract_header_info_from_cnum(cnum_path) if cnum_path else {'po_number': '', 'batch_number': '', 'sim_quantity': 0}
    po_number = cnum_info.get('po_number', '')
    batch_number = cnum_info.get('batch_number', '')
    batch_qty = cnum_info.get('sim_quantity', 0)
    
    # If CNUM didn't have the info, fallback to SCM
    if not po_number or not batch_number or batch_qty == 0:
        scm_po, scm_batch, scm_qty = extract_po_and_batch_from_scm(scm_path) if scm_path else ("", "", 0)
        if not po_number:
            po_number = scm_po
        if not batch_number:
            batch_number = scm_batch
        if batch_qty == 0:
            batch_qty = scm_qty
    
    # Determine final verification status based on validation errors
    error_count = len(validation_errors) if validation_errors else 0
    final_status = "OK" if error_count == 0 else "NOT OK"
    
    # Get current date in DD-MM-YYYY format
    verification_date = datetime.today().strftime('%d-%m-%Y')
    
    # Extract Perso Script filename from path
    perso_script_filename = os.path.basename(perso_script_path) if perso_script_path else ""
    
    # Batch info: combine batch number and quantity
    batch_info = f"{batch_number} ({batch_qty})" if batch_number else str(batch_qty)

    # Metadata labels and values
    metadata = [
        ("Operator Name", operator_name),
        ("SOF NO", sof_number),
        ("PO NO", po_number),
        ("Circle", circle_value),
        ("Chip", chip_code),
        ("Batch No & Batch QTY", batch_info),
        ("Total QTY", total_qty),
        ("Date of the verification", verification_date),
        ("Script used for Perso", perso_script_filename),
        ("Final Verification Status", final_status),
    ]

    for i, (label, value) in enumerate(metadata, start=4):  # Start from row 4
        # Label cell
        label_cell = ws.cell(row=i, column=1, value=label)
        label_cell.fill = styles['dark_blue_fill']
        label_cell.font = Font(bold=True, color='FFFFFF')
        label_cell.border = styles['thick_border']
        
        # Value cell - Apply bold formatting and left alignment to all header values
        value_cell = ws.cell(row=i, column=2, value=value)
        value_cell.border = styles['thick_border']
        value_cell.font = Font(bold=True)
        value_cell.alignment = Alignment(horizontal='left', vertical='center')

    return metadata

def update_final_verification_status(ws, validation_errors):
    """
    Update the Final Verification Status cell after validation is complete.
    This should be called after all validation errors have been collected.
    
    Parameters:
    - ws: Worksheet object
    - validation_errors: List of validation errors
    """
    # Find the row with "Final Verification Status" label
    for row in range(1, 20):  # Search in first 20 rows
        cell_value = ws.cell(row=row, column=1).value
        if cell_value == "Final Verification Status":
            # Calculate final status based on validation errors
            error_count = len(validation_errors) if validation_errors else 0
            final_status = "OK" if error_count == 0 else "NOT OK"
            
            # Update the value cell
            ws.cell(row=row, column=2, value=final_status)
            ws.cell(row=row, column=2).font = Font(bold=True)
            print(f"Updated Final Verification Status: {final_status} (errors: {error_count})")
            break
