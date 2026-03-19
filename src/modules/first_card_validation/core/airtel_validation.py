import os
import re
import traceback
import getpass
from difflib import SequenceMatcher
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Add at the top for image support
try:
    from openpyxl.drawing.image import Image as ExcelImage
    IMAGE_SUPPORT = True
    print("✅ Openpyxl image support available")
except ImportError:
    IMAGE_SUPPORT = False
    print("⚠️  Openpyxl image support not available. Images will not be inserted.")

# ========== HELPER FUNCTIONS ==========
def extract_sof_number(folder_path):
    """
    Extract SOF Number from folder name.
    If folder name contains RTLP prefix, use the complete RTLP code (e.g., RTLP10090).
    Folder names follow format like "RTLP10086 NBIOT RJ 50K" or "RTLP10090 - WBIoT BR 250K"
    """
    if not folder_path or not os.path.exists(folder_path):
        return ""
    
    folder_name = os.path.basename(folder_path)
    
    # Check for RTLP prefix
    rtlp_match = re.search(r'(RTLP\d+)', folder_name, re.IGNORECASE)
    if rtlp_match:
        return rtlp_match.group(1)
    
    return ""

# ========== EXCEL STYLING & HEADER FUNCTIONS ==========
def setup_excel_styles():
    """Setup Excel styles - UPDATED with exact colors from template"""
    return {
        'dark_blue_fill': PatternFill(start_color='BDDDE9', end_color='BDDDE9', fill_type='solid'),
        'header_font': Font(color='000000', bold=True, size=12),
        'thick_border': Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        ),
        'green_fill': PatternFill(start_color='BDDDE9', end_color='BDDDE9', fill_type='solid'),  # #BDDDE9
        'red_fill': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),    # #FF0000
        'yellow_fill': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'), # #FFFF00
        'gray_fill': PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),
    }

def setup_excel_headers(ws, styles, machine_log_path, pcom_path, cnum_path, cps_path):
    """
    Setup Excel headers - UPDATED: Removed Profile Type and File Paths sections
    """
    # Title Section
    ws["A1"] = "First Card Validation Report"
    ws["A1"].fill = styles['dark_blue_fill']
    ws["A1"].font = Font(bold=True, color='000000', size=14)
    ws["A1"].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells("A1:B1")
    
    # Timestamp in D1
    timestamp = datetime.today().strftime('%A, %d %B %Y %I:%M %p')
    ws["D1"] = f"Date : {timestamp}"
    ws["D1"].fill = styles['dark_blue_fill']
    ws["D1"].font = Font(bold=True, color='000000', size=14)
    ws["D1"].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells("D1:E1")
    
    # Username in D2
    username = getpass.getuser()
    ws["D2"] = f"User : {username}"
    ws["D2"].fill = styles['dark_blue_fill']
    ws["D2"].font = Font(bold=True, color='000000', size=14)
    ws["D2"].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells("D2:E2")
    
    ws.row_dimensions[1].height = 30
    
    # Metadata labels - AIRTEL specific
    pre_labels = [
        "Operator Name",
        "SOF NO", 
        "PO NO",
        "Circle",
        "Chip",
        "Batch No & Batch QTY",
        "Total QTY",
        "Date of the verification",
        "Script used for Perso",
        "Final Verification Status",
    ]
    
    # Write metadata labels with styling
    for i, label in enumerate(pre_labels, start=4):
        label_cell = ws.cell(row=i, column=1, value=label)
        label_cell.fill = styles['dark_blue_fill']
        label_cell.font = Font(bold=True, color='000000')
        label_cell.border = styles['thick_border']
        
        # Empty value cell with border
        ws.cell(row=i, column=2).border = styles['thick_border']
    
    # REMOVED: Profile Type section (rows 15-19)
    # REMOVED: File Paths section (rows 15-19)
    
    # Start validation table immediately after metadata section
    validation_row = 15  # Changed from 21 to 15 (after the 10 metadata rows + 1 empty row)
    
    # Add an empty row for spacing
    ws.row_dimensions[validation_row].height = 10
    
    return validation_row + 1  # Return the starting row for data table (header row)

# ========== IMAGE HANDLING FUNCTIONS ==========
def insert_image_to_excel(ws, image_path, cell_reference, width=150, height=100):
    """
    Insert an image into Excel at specified cell.
    """
    try:
        if not IMAGE_SUPPORT:
            print(f"⚠️  Image support not available. Skipping: {image_path}")
            return False
            
        if not image_path or not os.path.exists(image_path):
            print(f"⚠️  Image not found: {image_path}")
            return False
        
        img = ExcelImage(image_path)
        
        # Resize image
        img.width = width
        img.height = height
        
        # Add to worksheet
        ws.add_image(img, cell_reference)
        
        print(f"✅ Image inserted: {image_path} at {cell_reference}")
        return True
        
    except Exception as e:
        print(f"❌ Error inserting image: {str(e)}")
        return False

def add_image_section_to_report(ws, styles, image_paths, start_row):
    """
    Add image section to the report if images are provided.
    Returns: next available row after images
    """
    if not image_paths or not isinstance(image_paths, dict):
        print("📸 No images provided for AIRTEL report")
        return start_row
    
    try:
        print(f"📸 Processing AIRTEL images: {image_paths}")
        
        # Count valid images
        valid_images = {}
        for img_type, img_path in image_paths.items():
            if img_path and os.path.exists(img_path):
                valid_images[img_type] = img_path
        
        if not valid_images:
            print("📸 No valid images found (files don't exist)")
            return start_row
        
        # Create header for Images section
        img_header_row = start_row
        ws.cell(row=img_header_row, column=1, value="IMAGES").font = Font(size=14, bold=True)
        ws.merge_cells(f'A{img_header_row}:E{img_header_row}')
        ws.row_dimensions[img_header_row].height = 25
        
        img_row = img_header_row + 1
        
        # Process images based on type
        image_inserted = False
        
        if 'inner_label' in valid_images:
            inner_path = valid_images['inner_label']
            ws.cell(row=img_row, column=1, value="INNER LABEL")
            if insert_image_to_excel(ws, inner_path, f'B{img_row}'):
                ws.row_dimensions[img_row].height = 80  # Adjust row height for image
                print(f"✅ INNER LABEL image added: {inner_path}")
                image_inserted = True
            else:
                ws.cell(row=img_row, column=2, value=f"Image not inserted: {inner_path}")
                ws.cell(row=img_row, column=2).fill = styles['yellow_fill']
            img_row += 1
        
        if 'outer_label' in valid_images:
            outer_path = valid_images['outer_label']
            ws.cell(row=img_row, column=1, value="OUTER LABEL")
            if insert_image_to_excel(ws, outer_path, f'B{img_row}'):
                ws.row_dimensions[img_row].height = 80
                print(f"✅ OUTER LABEL image added: {outer_path}")
                image_inserted = True
            else:
                ws.cell(row=img_row, column=2, value=f"Image not inserted: {outer_path}")
                ws.cell(row=img_row, column=2).fill = styles['yellow_fill']
            img_row += 1
        
        # Add spacing after images
        img_row += 1
        
        if image_inserted:
            print(f"✅ Image section completed. Next row: {img_row}")
        else:
            print(f"⚠️  No images were successfully inserted")
            
        return img_row
        
    except Exception as e:
        print(f"❌ Error adding image section: {str(e)}")
        traceback.print_exc()
        return start_row

# ========== SAVE REPORT FUNCTION ==========
def save_report(wb, ml_path, pcom_path, iccid=None, sof_number=None):
    """
    Save workbook with AIRTEL naming convention using ICCID
    """
    try:
        # Create final report name using ICCID if available
        if iccid:
            # Clean ICCID for filename
            clean_iccid = str(iccid).replace("ICCID_", "").strip()
            # Include SOF number if available
            if sof_number:
                report_name = f"{clean_iccid}_{sof_number}_validation_report.xlsx"
            else:
                report_name = f"{clean_iccid}_validation_report.xlsx"
        else:
            # Fallback: Extract base filename
            base_name = os.path.splitext(os.path.basename(ml_path))[0]
            
            # Remove "Log_" prefix if present
            if base_name.lower().startswith("log_"):
                raw_number = base_name[4:]
            else:
                raw_number = base_name
            
            # Pairwise swap each 2 digits
            swapped_number = ""
            for i in range(0, len(raw_number), 2):
                if i + 1 < len(raw_number):
                    swapped_number += raw_number[i+1] + raw_number[i]
                else:
                    swapped_number += raw_number[i]
            
            # Create final report name
            # Include SOF number if available
            if sof_number:
                report_name = f"{swapped_number}_{sof_number}_validation_report.xlsx"
            else:
                report_name = f"{swapped_number}_validation_report.xlsx"
        
        # Save in PCOM folder if available, else Desktop
        if pcom_path and os.path.exists(os.path.dirname(pcom_path)):
            pcom_folder = os.path.dirname(pcom_path)
            report_path = os.path.join(pcom_folder, report_name)
        else:
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            report_path = os.path.join(desktop, report_name)
        
        wb.save(report_path)
        print(f"✅ Report saved successfully: {report_path}")
        return report_path
        
    except Exception as e:
        print(f"❌ Failed to save report: {e}")
        return None

# ========== VALIDATION RULES ==========
AIR_TEL_VALIDATION_RULES = {
    "PSK (6F2B)": ("from_value", "NR", "NR", "from_value"),
    "DEK1 (6F2B)": ("from_value", "NR", "NR", "from_value"),
    "IMSI": ("from_value", "from_value", "from_value", "from_value"),
    "PUK1": ("from_value", "from_value", "from_value", "NR"),
    "PUK2": ("from_value", "from_value", "from_value", "NR"),
    "ADM": ("from_value", "from_value", "NR", "NR"),
    "ICCID": ("from_value", "from_value", "from_value", "from_value"),
    "KIC1 (6F22)": ("from_value", "NR", "from_value", "from_value"),
    "KID1 (6F22)": ("from_value", "NR", "from_value", "from_value"),
    "KIK1 (6F22)": ("from_value", "NR", "NR", "from_value"),
    "KIC2 (6F22)": ("from_value", "NR", "NR", "from_value"),
    "KID2 (6F22)": ("from_value", "NR", "NR", "from_value"),
    "KIK2 (6F22)": ("from_value", "NR", "NR", "from_value"),
    "ACC": ("from_value", "from_value", "NR", "NR")
}

# ========== SIMILARITY CHECK FUNCTION ==========
def calculate_similarity(str1, str2):
    """
    Calculate similarity between two strings (0 to 1).
    Returns similarity percentage.
    """
    if not str1 or not str2:
        return 0.0
    
    # Use SequenceMatcher for better similarity calculation
    return SequenceMatcher(None, str(str1).strip().upper(), str(str2).strip().upper()).ratio() * 100

# ========== ICCID SWAPPING AND CONVERSION FUNCTIONS ==========
def swap_iccid(iccid):
    """
    Swap ICCID characters according to the specification.
    Format: Pairwise swapping of characters (0<->1, 2<->3, etc.)
    UPDATED: Handles alphanumeric characters (0-9, A-F, U)
    """
    if not iccid or len(iccid) < 2:
        return iccid
    
    swapped = ''
    for i in range(0, len(iccid), 2):
        if i + 1 < len(iccid):
            # Swap pairs of characters (digits or letters)
            swapped += iccid[i + 1] + iccid[i]
        else:
            swapped += iccid[i]  # Last character if odd length
    
    return swapped

def hex_to_ascii_iccid(hex_str):
    """
    Convert hex string to ASCII ICCID.
    Input: "39 38 31 39 30 31 32 33 34 35 36 37 38 39 30 39 38 30 46 30"
    Output: "981901234567890980F0"
    UPDATED: Preserves F/U characters correctly
    """
    if not hex_str:
        return ""
    
    # Remove spaces and convert hex to ASCII
    hex_chars = hex_str.replace(' ', '')
    ascii_str = ""
    
    for i in range(0, len(hex_chars), 2):
        if i + 2 <= len(hex_chars):
            hex_byte = hex_chars[i:i+2]
            try:
                # Convert hex to decimal
                decimal_val = int(hex_byte, 16)
                
                # Check if this represents an ASCII character
                if 32 <= decimal_val <= 126:  # Printable ASCII range
                    ascii_char = chr(decimal_val)
                    ascii_str += ascii_char
                else:
                    # If not printable ASCII, keep hex representation
                    ascii_str += hex_byte
            except:
                # If conversion fails, keep hex representation
                ascii_str += hex_byte
    
    return ascii_str

def hex_ascii_to_string(hex_ascii_str):
    """
    Convert hex ASCII string to readable string.
    Input: "3236333739323737" (hex ASCII representation)
    Output: "26379277" (ASCII characters)
    """
    if not hex_ascii_str:
        return ""
    
    # Remove spaces
    hex_ascii_str = hex_ascii_str.replace(' ', '')
    
    result = ""
    for i in range(0, len(hex_ascii_str), 2):
        if i + 2 <= len(hex_ascii_str):
            hex_pair = hex_ascii_str[i:i+2]
            try:
                # Convert hex to decimal
                dec_value = int(hex_pair, 16)
                # Convert to ASCII character
                ascii_char = chr(dec_value)
                result += ascii_char
            except:
                # If conversion fails, keep hex pair
                result += hex_pair
    
    return result

def iccid_for_cnum_comparison(ascii_iccid):
    """
    Prepare ICCID for CNUM comparison:
    1. Convert ASCII to hex if needed
    2. Swap the value
    3. Replace F with U
    UPDATED: Handles alphanumeric input directly
    """
    if not ascii_iccid:
        return ""
    
    # If it contains spaces, it's in hex format
    if ' ' in ascii_iccid:
        # Convert hex string to ASCII
        ascii_iccid = hex_to_ascii_iccid(ascii_iccid)
    elif re.match(r'^[0-9A-F]+$', ascii_iccid) and len(ascii_iccid) % 2 == 0:
        # If it's hex ASCII without spaces (like "393831393030...")
        ascii_iccid = hex_ascii_to_string(ascii_iccid)
    
    # Swap the ICCID
    swapped = swap_iccid(ascii_iccid)
    
    # Replace F with U for CNUM comparison
    swapped_for_cnum = swapped.replace('F', 'U').replace('f', 'U')
    
    return swapped_for_cnum

def iccid_for_cps_comparison(ascii_iccid):
    """
    Prepare ICCID for cps comparison:
    1. Convert ASCII to hex if needed
    2. Swap the value
    UPDATED: Handles alphanumeric input
    """
    if not ascii_iccid:
        return ""
    
    # Ensure we have the ASCII string
    if ' ' in ascii_iccid:
        ascii_iccid = hex_to_ascii_iccid(ascii_iccid)
    elif re.match(r'^[0-9A-F]+$', ascii_iccid) and len(ascii_iccid) % 2 == 0:
        # If it's hex ASCII without spaces
        ascii_iccid = hex_ascii_to_string(ascii_iccid)
    
    # Swap the ICCID
    swapped = swap_iccid(ascii_iccid)
    
    return swapped

def process_imsi_for_cnum_cps(imsi_18_digit):
    """
    Process IMSI for CNUM/cps comparison:
    1. Swap the 18-digit IMSI
    2. Remove first 3 digits to get 15-digit IMSI
    """
    if not imsi_18_digit or len(imsi_18_digit) != 18:
        return imsi_18_digit
    
    # Swap the IMSI
    swapped = swap_iccid(imsi_18_digit)
    
    # Remove first 3 digits to get 15-digit IMSI
    if len(swapped) >= 15:
        imsi_15_digit = swapped[3:]  # Remove first 3 digits
        return imsi_15_digit[:15]  # Ensure 15 digits
    
    return swapped

# ========== PARSING FUNCTIONS ==========
def extract_value(line, command):
    """Extract value from line after command - UPDATED for alphanumeric"""
    line = line.upper().strip()
    if command in line:
        parts = line.split(command)
        if len(parts) > 1:
            value = parts[1].split('SW9000')[0] if 'SW9000' in parts[1] else parts[1]
            # UPDATED: Allow alphanumeric characters (0-9, A-F)
            value = re.sub(r'[^0-9A-F]', '', value)
            return value
    return None

def parse_machine_log(filepath):
    """Parse Machine Log file - FIXED: Keep PUK as hex ASCII"""
    print("="*80)
    print("🚀 PARSING MACHINE LOG")
    print("="*80)
    
    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()
        
        extracted = {}
        
        for line_num, line in enumerate(lines, 1):
            line_upper = line.upper().strip()
            
            # IMSI - Keep existing (should be numeric)
            if '00D6000009' in line_upper:
                val = extract_value(line_upper, '00D6000009')
                if val:
                    print(f"🔍 Line {line_num}: Raw IMSI hex: {val}")
                    
                    # Convert hex to BCD digits
                    decimal_digits = []
                    for i in range(0, len(val), 2):
                        if i + 2 <= len(val):
                            hex_byte = val[i:i+2]
                            try:
                                byte_value = int(hex_byte, 16)
                                high_digit = (byte_value >> 4) & 0x0F
                                low_digit = byte_value & 0x0F
                                
                                # Check if valid BCD digit (0-9) for IMSI
                                if 0 <= high_digit <= 9:
                                    decimal_digits.append(str(high_digit))
                                if 0 <= low_digit <= 9:
                                    decimal_digits.append(str(low_digit))
                            except:
                                continue
                    
                    if decimal_digits:
                        if len(decimal_digits) >= 18:
                            imsi_18_digit = ''.join(decimal_digits)[:18]
                            extracted['IMSI'] = imsi_18_digit
                            print(f"✅ IMSI parsed (18 digits): {imsi_18_digit}")
                        else:
                            imsi = ''.join(decimal_digits)
                            extracted['IMSI'] = imsi
                            print(f"✅ IMSI parsed: {imsi}")
            
            # ICCID - UPDATED: Extract alphanumeric (0-9, A-F)
            elif '00D600000A' in line_upper:
                val = extract_value(line_upper, '00D600000A')
                if val:
                    print(f"🔍 Line {line_num}: Raw ICCID hex: {val}")
                    
                    # Take first 20 hex chars (10 bytes) for ICCID
                    hex_str = val[:20].ljust(20, '0')  # Pad to 20 chars if shorter
                    
                    # UPDATED: Convert hex to alphanumeric ICCID
                    iccid_chars = []
                    for i in range(0, len(hex_str), 2):
                        if i + 2 <= len(hex_str):
                            hex_byte = hex_str[i:i+2]
                            try:
                                byte_value = int(hex_byte, 16)
                                
                                # Convert each nibble to character
                                high_nibble = (byte_value >> 4) & 0x0F
                                low_nibble = byte_value & 0x0F
                                
                                # UPDATED: Handle values 0-15 (0-9, A-F)
                                def nibble_to_char(nibble):
                                    if 0 <= nibble <= 9:
                                        return str(nibble)
                                    elif 10 <= nibble <= 15:
                                        return chr(ord('A') + nibble - 10)
                                    return ''
                                
                                high_char = nibble_to_char(high_nibble)
                                low_char = nibble_to_char(low_nibble)
                                
                                if high_char:
                                    iccid_chars.append(high_char)
                                if low_char:
                                    iccid_chars.append(low_char)
                                    
                            except:
                                continue
                    
                    if iccid_chars:
                        # Extract alphanumeric ICCID (18-20 characters)
                        iccid = ''.join(iccid_chars)[:20]  # Can be up to 20 chars
                        extracted['ICCID'] = iccid
                        print(f"✅ ICCID parsed (alphanumeric): {iccid}")
            
            # PSK/DEK1
            elif '00D600002AFE85400310' in line_upper and 'FE80400210' in line_upper:
                parts = line_upper.split('00D600002AFE85400310')
                if len(parts) > 1:
                    tail = parts[1]
                    if 'FE80400210' in tail:
                        psk = tail.split('FE80400210')[0][:32]
                        dek1 = tail.split('FE80400210')[1][:32]
                        extracted['PSK (6F2B)'] = psk
                        extracted['DEK1 (6F2B)'] = dek1
                        print(f"✅ PSK parsed: {psk}")
                        print(f"✅ DEK1 parsed: {dek1}")
            
            # PUK1 - FIXED: Keep as hex ASCII (3236333739323737)
            elif '00D6000015F00303' in line_upper:
                val = extract_value(line_upper, '00D6000015F00303')
                if val:
                    print(f"🔍 Line {line_num}: Raw PUK1 hex: {val}")
                    if 'FFFFFFFF0A0A' in val:
                        parts = val.split('FFFFFFFF0A0A')
                        if len(parts) > 1 and len(parts[1]) >= 16:
                            hex_ascii_puk = parts[1][:16]  # This is 3236333739323737
                            # Store the hex ASCII value AS-IS
                            extracted['PUK1'] = hex_ascii_puk
                            print(f"✅ PUK1 parsed (hex ASCII): {hex_ascii_puk}")
                    else:
                        # Try to extract 16 hex chars
                        hex_match = re.search(r'([0-9A-F]{16})', val)
                        if hex_match:
                            hex_ascii_puk = hex_match.group(1)
                            extracted['PUK1'] = hex_ascii_puk
                            print(f"✅ PUK1 parsed (direct): {hex_ascii_puk}")
            
            # PUK2 - FIXED: Keep as hex ASCII
            elif '00D6000015E00303' in line_upper:
                val = extract_value(line_upper, '00D6000015E00303')
                if val:
                    print(f"🔍 Line {line_num}: Raw PUK2 hex: {val}")
                    if 'FFFFFFFF0A0A' in val:
                        parts = val.split('FFFFFFFF0A0A')
                        if len(parts) > 1 and len(parts[1]) >= 16:
                            hex_ascii_puk = parts[1][:16]
                            # Store the hex ASCII value AS-IS
                            extracted['PUK2'] = hex_ascii_puk
                            print(f"✅ PUK2 parsed (hex ASCII): {hex_ascii_puk}")
                    else:
                        # Try to extract 16 hex chars
                        hex_match = re.search(r'([0-9A-F]{16})', val)
                        if hex_match:
                            hex_ascii_puk = hex_match.group(1)
                            extracted['PUK2'] = hex_ascii_puk
                            print(f"✅ PUK2 parsed (direct): {hex_ascii_puk}")
            
            # ADM
            elif '00D600000B800A0A' in line_upper:
                val = extract_value(line_upper, '00D600000B800A0A')
                if val and len(val) >= 16:
                    extracted['ADM'] = val[:16]
                    print(f"✅ ADM parsed: {val[:16]}")
            
            # ACC
            elif '00D6000002' in line_upper:
                val = extract_value(line_upper, '00D6000002')
                if val and len(val) >= 4:
                    extracted['ACC'] = val[:4]
                    print(f"✅ ACC parsed: {val[:4]}")
            
            # KIC1
            elif '00DC01041BFE0110' in line_upper:
                val = extract_value(line_upper, '00DC01041BFE0110')
                if val:
                    val = val.replace('FFFFFFFFFFFFFFFF', '')
                    if len(val) >= 32:
                        extracted['KIC1 (6F22)'] = val[:32]
                        print(f"✅ KIC1 parsed: {val[:32]}")
            
            # KID1
            elif '00DC02041BFE0111' in line_upper:
                val = extract_value(line_upper, '00DC02041BFE0111')
                if val:
                    val = val.replace('FFFFFFFFFFFFFFFF', '')
                    if len(val) >= 32:
                        extracted['KID1 (6F22)'] = val[:32]
                        print(f"✅ KID1 parsed: {val[:32]}")
            
            # KIK1
            elif '00DC03041BFE0112' in line_upper:
                val = extract_value(line_upper, '00DC03041BFE0112')
                if val:
                    val = val.replace('FFFFFFFFFFFFFFFF', '')
                    if len(val) >= 32:
                        extracted['KIK1 (6F22)'] = val[:32]
                        print(f"✅ KIK1 parsed: {val[:32]}")
            
            # KIC2
            elif '00DC04041BFE0050' in line_upper:
                val = extract_value(line_upper, '00DC04041BFE0050')
                if val:
                    val = val.replace('FFFFFFFFFFFFFFFF', '')
                    if len(val) >= 32:
                        extracted['KIC2 (6F22)'] = val[:32]
                        print(f"✅ KIC2 parsed: {val[:32]}")
            
            # KID2
            elif '00DC05041BFE0051' in line_upper:
                val = extract_value(line_upper, '00DC05041BFE0051')
                if val:
                    val = val.replace('FFFFFFFFFFFFFFFF', '')
                    if len(val) >= 32:
                        extracted['KID2 (6F22)'] = val[:32]
                        print(f"✅ KID2 parsed: {val[:32]}")
            
            # KIK2
            elif '00DC06041BFE0052' in line_upper:
                val = extract_value(line_upper, '00DC06041BFE0052')
                if val:
                    val = val.replace('FFFFFFFFFFFFFFFF', '')
                    if len(val) >= 32:
                        extracted['KIK2 (6F22)'] = val[:32]
                        print(f"✅ KIK2 parsed: {val[:32]}")
        
        # Ensure all fields exist
        for field in AIR_TEL_VALIDATION_RULES.keys():
            if field not in extracted:
                extracted[field] = "Not Found"
        
        found_count = sum(1 for v in extracted.values() if v != "Not Found")
        print(f"✅ Extracted {found_count}/{len(AIR_TEL_VALIDATION_RULES)} fields from Machine Log")
        
        return extracted
        
    except Exception as e:
        print(f"❌ Machine Log parsing error: {str(e)}")
        traceback.print_exc()
        return {}

def parse_pcom_file(pcom_path):
    """Parse PCOM file - UPDATED for alphanumeric ICCID"""
    if not pcom_path or not os.path.exists(pcom_path):
        print("⚠️  PCOM file not provided or not found")
        return {}
    
    try:
        print(f"\n📋 PARSING PCOM FILE: {pcom_path}")
        with open(pcom_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        
        fields = {}
        
        # Try different patterns for each field
        patterns = [
            ('IMSI', r'\.DEFINE\s+%IMSI\s+"?([0-9]+)"?'),
            ('IMSI', r'IMSI\s*=\s*"?([0-9]+)"?'),
            ('ICCID', r'\.DEFINE\s+%ICCID\s+"?([0-9A-F]{18,20})"?'),  # UPDATED: alphanumeric
            ('ICCID', r'ICCID\s*=\s*"?([0-9A-F]{18,20})"?'),  # UPDATED: alphanumeric
            ('PUK1', r'\.DEFINE\s+%PUK1\s+"?([0-9A-F]{16})"?'),
            ('PUK1', r'PUK1\s*=\s*"?([0-9A-F]{16})"?'),
            ('PUK2', r'\.DEFINE\s+%PUK2\s+"?([0-9A-F]{16})"?'),  # Added for PUK2
            ('PUK2', r'PUK2\s*=\s*"?([0-9A-F]{16})"?'),  # Added for PUK2
            ('ADM', r'\.DEFINE\s+%ISC1\s+"?([0-9A-F]{16})"?'),
            ('ADM', r'ADM\s*=\s*"?([0-9A-F]{16})"?'),
            ('KIC1 (6F22)', r'\.DEFINE\s+%KIC1\s+"?([0-9A-F]{32})"?'),
            ('KIC1 (6F22)', r'KIC1\s*=\s*"?([0-9A-F]{32})"?'),
            ('KID1 (6F22)', r'\.DEFINE\s+%KID1\s+"?([0-9A-F]{32})"?'),
            ('KID1 (6F22)', r'KID1\s*=\s*"?([0-9A-F]{32})"?'),
            ('ACC', r'\.DEFINE\s+%ACC\s+"?([0-9]{4})"?'),
            ('ACC', r'ACC\s*=\s*"?([0-9]{4})"?'),
        ]
        
        for field_name, pattern in patterns:
            if field_name not in fields:  # Only find if not already found
                match = re.search(pattern, content, re.IGNORECASE)
                if match:
                    value = match.group(1)
                    fields[field_name] = value
                    print(f"✅ PCOM {field_name}: {value}")
        
        print(f"✅ PCOM parsed: {len(fields)} fields found")
        return fields
        
    except Exception as e:
        print(f"❌ PCOM parsing error: {str(e)}")
        traceback.print_exc()
        return {}

def parse_cnum_file(cnum_path):
    """Parse CNUM file - FIXED for PUK comparison"""
    if not cnum_path or not os.path.exists(cnum_path):
        print("⚠️  CNUM file not provided or not found")
        return {}
    
    try:
        print(f"\n📋 PARSING CNUM FILE: {cnum_path}")
        with open(cnum_path, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()
        
        fields = {}
        
        print(f"📊 Total lines in CNUM: {len(lines)}")
        
        # Look for specific structure: line 24 has headers, line 25 has data
        if len(lines) >= 25:
            # Line 24 (index 23): Header line
            header_line = lines[23].strip().upper() if len(lines) > 23 else ""
            # Line 25 (index 24): Data line
            data_line = lines[24].strip() if len(lines) > 24 else ""
            
            print(f"📋 Line 24 (header): {header_line[:100]}...")
            print(f"📋 Line 25 (data): {data_line[:100]}...")
            
            # Check if this is the correct structure
            if 'VAR_OUT:' in header_line and data_line:
                # Parse headers
                headers_part = header_line.replace('VAR_OUT:', '').strip()
                header_parts = [h.strip() for h in headers_part.split('/') if h.strip()]
                
                print(f"📋 CNUM Headers ({len(header_parts)}):")
                for i, header in enumerate(header_parts):
                    print(f"   {i+1}. {header}")
                
                # Parse data
                data_parts = data_line.split()
                
                print(f"📋 CNUM Data ({len(data_parts)} parts):")
                for i, data in enumerate(data_parts):
                    print(f"   {i+1}. {data}")
                
                # Map headers to data
                for i, header in enumerate(header_parts):
                    if i < len(data_parts):
                        data_value = data_parts[i]
                        
                        # Map based on header name
                        if header == 'IMSI':
                            digits = re.sub(r'[^0-9]', '', data_value)
                            if len(digits) >= 15:
                                fields['IMSI'] = digits[:15]
                                print(f"✅ CNUM IMSI: {digits[:15]}")
                            else:
                                fields['IMSI'] = data_value
                                print(f"⚠️  CNUM IMSI (raw): {data_value}")
                        
                        elif header == 'ICCID':
                            # UPDATED: Allow alphanumeric (0-9, A-F, U)
                            cleaned = re.sub(r'[^0-9A-FU]', '', data_value.upper())
                            if len(cleaned) >= 18:
                                fields['ICCID'] = cleaned[:20]  # Up to 20 chars
                                print(f"✅ CNUM ICCID (alphanumeric): {cleaned[:20]}")
                            elif len(data_value) >= 18:
                                fields['ICCID'] = data_value[:20]
                                print(f"✅ CNUM ICCID (raw): {data_value[:20]}")
                        
                        elif header == 'PUK1':
                            # CNUM has ASCII string, not hex ASCII
                            fields['PUK1'] = data_value
                            print(f"✅ CNUM PUK1: {data_value}")
                        
                        elif header == 'PUK2':
                            fields['PUK2'] = data_value
                            print(f"✅ CNUM PUK2: {data_value}")
                        
                        elif header == 'CIPHERKEY_RFM':
                            hex_match = re.search(r'([0-9A-F]{32})', data_value, re.IGNORECASE)
                            if hex_match:
                                fields['KIC1 (6F22)'] = hex_match.group(1)
                                print(f"✅ CNUM KIC1 (CIPHERKEY_RFM): {hex_match.group(1)}")
                            elif len(data_value) >= 32:
                                fields['KIC1 (6F22)'] = data_value[:32]
                                print(f"✅ CNUM KIC1 (raw): {data_value[:32]}")
                        
                        elif header == 'MACKEY_RFM':
                            hex_match = re.search(r'([0-9A-F]{32})', data_value, re.IGNORECASE)
                            if hex_match:
                                fields['KID1 (6F22)'] = hex_match.group(1)
                                print(f"✅ CNUM KID1 (MACKEY_RFM): {hex_match.group(1)}")
                            elif len(data_value) >= 32:
                                fields['KID1 (6F22)'] = data_value[:32]
                                print(f"✅ CNUM KID1 (raw): {data_value[:32]}")
                        
                        if header in ['A4IND', 'ACC'] and data_value.isdigit() and len(data_value) == 4:
                            fields['ACC'] = data_value
                            print(f"✅ CNUM ACC: {data_value}")
                
                # Also search entire data line for patterns
                for data_value in data_parts:
                    if len(data_value) == 4 and data_value.isdigit() and 'ACC' not in fields:
                        fields['ACC'] = data_value
                        print(f"✅ CNUM ACC (found in data): {data_value}")
                    
                    # Look for ICCID with U/F character
                    if 'ICCID' not in fields and len(data_value) >= 18:
                        if 'U' in data_value.upper() or data_value[-1].upper() in ['U', 'F']:
                            cleaned = re.sub(r'[^0-9A-FU]', '', data_value.upper())
                            if len(cleaned) >= 18:
                                fields['ICCID'] = cleaned[:20]
                                print(f"✅ CNUM ICCID (U/F pattern): {cleaned[:20]}")
                    
                    # Look for KIC1/KID1 if not found yet
                    if 'KIC1 (6F22)' not in fields:
                        hex_match = re.search(r'([0-9A-F]{32})', data_value, re.IGNORECASE)
                        if hex_match:
                            fields['KIC1 (6F22)'] = hex_match.group(1)
                            print(f"✅ CNUM KIC1 (pattern): {hex_match.group(1)}")
                    
                    if 'KID1 (6F22)' not in fields:
                        hex_match = re.search(r'([0-9A-F]{32})', data_value, re.IGNORECASE)
                        if hex_match and hex_match.group(1) != fields.get('KIC1 (6F22)', ''):
                            fields['KID1 (6F22)'] = hex_match.group(1)
                            print(f"✅ CNUM KID1 (pattern): {hex_match.group(1)}")
        
        # If standard parsing didn't work, search entire file
        if len(fields) < 3:
            print("⚠️  Using pattern-based CNUM parsing...")
            all_content = ' '.join([line.strip() for line in lines])
            
            # Extract IMSI (15 digits)
            imsi_matches = re.findall(r'\b(\d{15})\b', all_content)
            if imsi_matches and 'IMSI' not in fields:
                fields['IMSI'] = imsi_matches[0]
                print(f"✅ CNUM IMSI (pattern): {imsi_matches[0]}")
            
            # UPDATED: Extract ICCID (18-20 alphanumeric chars, may contain U/F)
            iccid_pattern = r'\b([0-9A-FU]{18,20})\b'
            iccid_matches = re.findall(iccid_pattern, all_content, re.IGNORECASE)
            for match in iccid_matches:
                if 'IMSI' not in fields or match != fields.get('IMSI', ''):
                    if match.startswith('89') or 'U' in match.upper() or match[-1].upper() in ['U', 'F']:
                        fields['ICCID'] = match.upper()[:20]
                        print(f"✅ CNUM ICCID (alphanumeric pattern): {match.upper()[:20]}")
                        break
            
            # Extract PUK1/PUK2 (8+ digits)
            puk_matches = re.findall(r'\b(\d{8,})\b', all_content)
            for match in puk_matches:
                if match != fields.get('IMSI', '') and match != fields.get('ICCID', ''):
                    if 'PUK1' not in fields:
                        fields['PUK1'] = match
                        print(f"✅ CNUM PUK1 (pattern): {match}")
                    elif 'PUK2' not in fields:
                        fields['PUK2'] = match
                        print(f"✅ CNUM PUK2 (pattern): {match}")
                        break
            
            # Extract KIC1/KID1 (32 hex)
            hex_32_matches = re.findall(r'\b([0-9A-F]{32})\b', all_content, re.IGNORECASE)
            if hex_32_matches:
                if 'KIC1 (6F22)' not in fields and len(hex_32_matches) > 0:
                    fields['KIC1 (6F22)'] = hex_32_matches[0]
                    print(f"✅ CNUM KIC1 (hex pattern): {hex_32_matches[0]}")
                if 'KID1 (6F22)' not in fields and len(hex_32_matches) > 1:
                    fields['KID1 (6F22)'] = hex_32_matches[1]
                    print(f"✅ CNUM KID1 (hex pattern): {hex_32_matches[1]}")
            
            # Extract ACC (4 digits)
            acc_matches = re.findall(r'\b(\d{4})\b', all_content)
            if acc_matches and 'ACC' not in fields:
                fields['ACC'] = acc_matches[0]
                print(f"✅ CNUM ACC (pattern): {acc_matches[0]}")
        
        print(f"✅ CNUM parsed: {len(fields)} fields found")
        for field, value in fields.items():
            print(f"   - {field}: {value}")
        
        return fields
        
    except Exception as e:
        print(f"❌ CNUM parsing error: {str(e)}")
        traceback.print_exc()
        return {}

def parse_cps_file(cps_path, machine_log_values=None):
    """
    Parse cps file with 80% similarity validation.
    Returns only values that are at least 80% similar to ML values.
    """
    if not cps_path or not os.path.exists(cps_path):
        print("⚠️  cps file not provided or not found")
        return {}
    
    try:
        print(f"\n📋 PARSING cps FILE (with 80% similarity check): {cps_path}")
        
        # Read cps content
        with open(cps_path, 'r', encoding='utf-8', errors='ignore') as f:
            cps_content = f.read()
        
        print(f"📊 cps file size: {len(cps_content)} characters")
        
        fields = {}
        
        # If we have Machine Log values, search for them with similarity check
        if machine_log_values:
            print("\n🔍 Searching for Machine Log values in cps (80% similarity)...")
            
            for field_name, ml_value in machine_log_values.items():
                if ml_value == "Not Found" or not ml_value:
                    continue
                
                ml_str = str(ml_value).strip()
                if len(ml_str) < 4:  # Skip very short values
                    continue
                
                print(f"\n   Checking {field_name}:")
                print(f"     ML value: {ml_str}")
                
                # Skip fields that are Not Required for cps
                if field_name in ["PUK1", "PUK2", "ADM", "ACC"]:
                    print(f"     Skipping - Not required for cps validation")
                    continue
                
                # Search for the value in cps
                ml_upper = ml_str.upper()
                cps_upper = cps_content.upper()
                
                # Try to find exact match first
                exact_pos = cps_upper.find(ml_upper)
                if exact_pos != -1:
                    # Extract the exact value
                    cps_value = cps_content[exact_pos:exact_pos+len(ml_str)]
                    fields[field_name] = cps_value
                    print(f"     ✅ Exact match found: {cps_value}")
                    continue
                
                # If no exact match, search for similar patterns
                print(f"     No exact match found, searching for similar patterns...")
                
                # For ICCID: look for alphanumeric patterns (18-20 chars)
                if field_name == "ICCID":
                    iccid_patterns = [
                        r'\b([0-9A-F]{18,20})\b',
                        r'ICCID[:\s]+([0-9A-F]{18,20})',
                        r'"([0-9A-F]{18,20})"',
                    ]
                    
                    best_match = None
                    best_similarity = 0
                    
                    for pattern in iccid_patterns:
                        matches = re.findall(pattern, cps_content, re.IGNORECASE)
                        for match in matches:
                            similarity = calculate_similarity(ml_str, match)
                            print(f"     Found: {match} (Similarity: {similarity:.1f}%)")
                            
                            if similarity > best_similarity:
                                best_similarity = similarity
                                best_match = match
                    
                    if best_match and best_similarity >= 80.0:
                        fields[field_name] = best_match
                        print(f"     ✅ Accepted: {best_match} ({best_similarity:.1f}% similar)")
                    elif best_match:
                        print(f"     ❌ Rejected: {best_match} ({best_similarity:.1f}% < 80%)")
                    else:
                        print(f"     ❌ No ICCID pattern found")
                
                # For IMSI: look for numeric patterns (15 or 18 digits)
                elif field_name == "IMSI":
                    # Try multiple patterns for IMSI
                    imsi_patterns = [
                        r'\b(\d{15})\b',      # 15-digit IMSI
                        r'\b(\d{18})\b',      # 18-digit IMSI
                        r'IMSI[:\s]+(\d{15,18})',
                        r'"(\d{15,18})"',
                    ]
                    
                    best_match = None
                    best_similarity = 0
                    best_match_info = ""
                    
                    for pattern in imsi_patterns:
                        matches = re.findall(pattern, cps_content)
                        for match in matches:
                            # For IMSI comparison, we need to handle both 15 and 18 digit formats
                            ml_digits = re.sub(r'\D', '', ml_str)
                            cps_digits = re.sub(r'\D', '', match)
                            
                            # Calculate similarity on numeric digits only
                            similarity = calculate_similarity(ml_digits, cps_digits)
                            print(f"     Found: {match} (Similarity: {similarity:.1f}%)")
                            
                            if similarity > best_similarity:
                                best_similarity = similarity
                                best_match = match
                                best_match_info = f"Found: {match}"
                    
                    # Check if we found a match with sufficient similarity
                    if best_match and best_similarity >= 80.0:
                        fields[field_name] = best_match
                        print(f"     ✅ Accepted: {best_match} ({best_similarity:.1f}% similar)")
                        
                        # Special handling for IMSI: Show the actual mismatch
                        ml_digits = re.sub(r'\D', '', ml_str)
                        cps_digits = re.sub(r'\D', '', best_match)
                        
                        if ml_digits != cps_digits:
                            print(f"     ⚠️  Note: IMSI mismatch detected: ML={ml_digits}, cps={cps_digits}")
                            # Store the mismatch info for later comparison
                            fields[f"{field_name}_cps_RAW"] = best_match
                    elif best_match:
                        print(f"     ❌ Rejected: {best_match} ({best_similarity:.1f}% < 80%)")
                    else:
                        print(f"     ❌ No IMSI pattern found")
                
                # For hex keys (PSK, DEK1, KIC1, etc.): look for 32-char hex patterns
                elif field_name in ["PSK (6F2B)", "DEK1 (6F2B)", "KIC1 (6F22)", "KID1 (6F22)", 
                                   "KIK1 (6F22)", "KIC2 (6F22)", "KID2 (6F22)", "KIK2 (6F22)"]:
                    hex_pattern = r'\b([0-9A-F]{32})\b'
                    matches = re.findall(hex_pattern, cps_content, re.IGNORECASE)
                    
                    best_match = None
                    best_similarity = 0
                    
                    for match in matches:
                        similarity = calculate_similarity(ml_str, match)
                        print(f"     Found: {match} (Similarity: {similarity:.1f}%)")
                        
                        if similarity > best_similarity:
                            best_similarity = similarity
                            best_match = match
                    
                    if best_match and best_similarity >= 80.0:
                        fields[field_name] = best_match
                        print(f"     ✅ Accepted: {best_match} ({best_similarity:.1f}% similar)")
                    elif best_match:
                        print(f"     ❌ Rejected: {best_match} ({best_similarity:.1f}% < 80%)")
                    else:
                        print(f"     ❌ No hex pattern found")
        
        print(f"\n✅ cps parsing completed: {len(fields)} values found (≥80% similarity)")
        if fields:
            print("📋 ACCEPTED values:")
            for field, value in fields.items():
                if not field.endswith("_cps_RAW"):  # Skip raw mismatch info fields
                    print(f"   - {field}: {value}")
        else:
            print("📋 No values found with ≥80% similarity")
        
        return fields
        
    except Exception as e:
        print(f"❌ cps parsing error: {str(e)}")
        traceback.print_exc()
        return {}

# ========== COMPARISON FUNCTIONS ==========
def compare_iccid(ml_value, external_value, comparison_type="PCOM", field_name="ICCID"):
    """
    Compare ICCID values with special handling.
    Returns: (is_match, error_message)
    """
    if not ml_value or ml_value == "Not Found":
        return False, f"{field_name}: Missing in Machine Log"
    
    if not external_value or external_value == "Not Found":
        return False, f"{field_name}: Missing in {comparison_type}"
    
    ml_value_str = str(ml_value).strip().upper()
    external_value_str = str(external_value).strip().upper()
    
    print(f"\n🔍 COMPARING ICCID:")
    print(f"   Comparison Type: {comparison_type}")
    print(f"   ML Value: {ml_value_str}")
    print(f"   External Value: {external_value_str}")
    
    # For cps comparison
    if comparison_type == "cps":
        # First check similarity (even though parse_cps_file already did)
        similarity = calculate_similarity(ml_value_str, external_value_str)
        if similarity < 80.0:
            error_msg = f"{field_name}: Low similarity in cps ({similarity:.1f}% < 80%) - Expected: {ml_value_str}, Got: {external_value_str}"
            return False, error_msg
        
        # Prepare ML ICCID for cps comparison
        ml_swapped = swap_iccid(ml_value_str)
        print(f"   ML swapped for cps: {ml_swapped}")
        
        # Case 1: cps has swapped ICCID (should match ml_swapped)
        if ml_swapped == external_value_str:
            print(f"   ✅ Match: ML swapped == cps value")
            return True, ""
        
        # Case 2: cps has original ICCID (should match ml_value_str)
        if ml_value_str == external_value_str:
            print(f"   ✅ Match: ML original == cps value")
            return True, ""
        
        # Case 3: Try swapping cps value too
        cps_swapped = swap_iccid(external_value_str)
        if ml_value_str == cps_swapped:
            print(f"   ✅ Match: ML original == cps swapped")
            return True, ""
        
        error_msg = f"{field_name}: Mismatch in {comparison_type} - Expected: {ml_value_str}, Got: {external_value_str}"
        print(f"   ❌ No match found")
        return False, error_msg
    
    # For CNUM comparison
    elif comparison_type == "CNUM":
        # Prepare ML ICCID for CNUM comparison
        ml_swapped = swap_iccid(ml_value_str)
        ml_for_cnum = ml_swapped.replace('F', 'U')
        
        print(f"   ML processed for CNUM: {ml_for_cnum}")
        print(f"   CNUM value: {external_value_str}")
        
        if ml_for_cnum == external_value_str:
            return True, ""
        
        error_msg = f"{field_name}: Mismatch in {comparison_type} - Expected (ML processed): {ml_for_cnum}, Got: {external_value_str}"
        return False, error_msg
    
    # For PCOM comparison (direct)
    else:
        if ml_value_str == external_value_str:
            return True, ""
        
        error_msg = f"{field_name}: Mismatch in {comparison_type} - Expected: {ml_value_str}, Got: {external_value_str}"
        return False, error_msg

def compare_imsi(ml_value, external_value, comparison_type="PCOM", field_name="IMSI"):
    """
    Compare IMSI values.
    Returns: (is_match, error_message)
    """
    if not ml_value or ml_value == "Not Found":
        return False, f"{field_name}: Missing in Machine Log"
    
    if not external_value or external_value == "Not Found":
        return False, f"{field_name}: Missing in {comparison_type}"
    
    ml_value_str = str(ml_value).strip()
    external_value_str = str(external_value).strip()
    
    print(f"\n🔍 COMPARING IMSI ({comparison_type}):")
    print(f"   ML IMSI: {ml_value_str}")
    print(f"   External IMSI: {external_value_str}")
    
    # Extract digits only for comparison
    ml_digits = re.sub(r'\D', '', ml_value_str)
    external_digits = re.sub(r'\D', '', external_value_str)
    
    print(f"   ML digits: {ml_digits}")
    print(f"   External digits: {external_digits}")
    
    if comparison_type == "CNUM":
        # For CNUM, process 18-digit ML IMSI to 15-digit
        if len(ml_digits) == 18:
            ml_processed = process_imsi_for_cnum_cps(ml_digits)
            print(f"   ML processed for CNUM: {ml_processed}")
            if ml_processed == external_digits[:15]:
                return True, ""
            
            error_msg = f"{field_name}: Mismatch in {comparison_type} - Expected (ML processed): {ml_processed}, Got: {external_digits[:15]}"
            return False, error_msg
        else:
            # Direct 15-digit comparison
            if ml_digits[:15] == external_digits[:15]:
                return True, ""
            
            error_msg = f"{field_name}: Mismatch in {comparison_type} - Expected: {ml_digits[:15]}, Got: {external_digits[:15]}"
            return False, error_msg
    
    elif comparison_type == "cps":
        # For cps, we already did similarity check in parse_cps_file
        # Now do detailed comparison
        
        # First, check if they're exactly the same
        if ml_digits == external_digits:
            print(f"   ✅ Exact match")
            return True, ""
        
        # Check if one is 18-digit and the other is 15-digit
        if len(ml_digits) == 18 and len(external_digits) == 15:
            # Check if cps IMSI is part of ML IMSI
            if external_digits in ml_digits:
                print(f"   ✅ cps IMSI found within ML IMSI")
                return True, ""
            
            # Also check processed ML IMSI
            ml_processed = process_imsi_for_cnum_cps(ml_digits)
            if ml_processed == external_digits:
                print(f"   ✅ Match with processed ML IMSI")
                return True, ""
        
        # Check if both are 18-digit but slightly different
        if len(ml_digits) == 18 and len(external_digits) == 18:
            # Calculate how many digits are different
            diff_count = sum(1 for a, b in zip(ml_digits, external_digits) if a != b)
            similarity = ((18 - diff_count) / 18) * 100
            
            print(f"   Both 18-digit IMSIs, {diff_count} digits different ({similarity:.1f}% similar)")
            
            # Even with high similarity, it's still a mismatch
            error_msg = f"{field_name}: Incorrect data in cps - Expected: {ml_digits}, Got: {external_digits} ({diff_count} digit(s) different)"
            return False, error_msg
        
        # Check if ML is 15-digit and cps is 18-digit
        if len(ml_digits) == 15 and len(external_digits) == 18:
            # Check if ML IMSI is part of cps IMSI
            if ml_digits in external_digits:
                print(f"   ✅ ML IMSI found within cps IMSI")
                return True, ""
        
        # Generic mismatch
        error_msg = f"{field_name}: Incorrect data in cps - Expected: {ml_digits}, Got: {external_digits}"
        print(f"   ❌ No match found")
        return False, error_msg
    
    # For PCOM comparison (direct)
    else:
        if ml_digits == external_digits:
            return True, ""
        
        error_msg = f"{field_name}: Mismatch in {comparison_type} - Expected: {ml_digits}, Got: {external_digits}"
        return False, error_msg

def compare_puk(ml_value, external_value, comparison_type="PCOM", field_name="PUK"):
    """
    Compare PUK values.
    Returns: (is_match, error_message)
    """
    if not ml_value or ml_value == "Not Found":
        return False, f"{field_name}: Missing in Machine Log"
    
    if not external_value or external_value == "Not Found":
        return False, f"{field_name}: Missing in {comparison_type}"
    
    ml_str = str(ml_value).strip()
    external_str = str(external_value).strip()
    
    print(f"\n🔍 PUK COMPARISON ({comparison_type}):")
    print(f"   ML: {ml_str}")
    print(f"   External: {external_str}")
    
    if comparison_type == "PCOM":
        # Direct hex ASCII comparison
        if ml_str == external_str:
            print(f"   ✅ Direct hex ASCII comparison")
            return True, ""
        
        error_msg = f"{field_name}: Mismatch in {comparison_type} - Expected: {ml_str}, Got: {external_str}"
        return False, error_msg
    
    elif comparison_type == "CNUM":
        # Convert ML hex ASCII to string
        ml_converted = hex_ascii_to_string(ml_str)
        print(f"   ML converted: {ml_str} -> {ml_converted}")
        
        if ml_converted == external_str:
            print(f"   ✅ Compare with CNUM")
            return True, ""
        
        error_msg = f"{field_name}: Mismatch in {comparison_type} - Expected (ML converted): {ml_converted}, Got: {external_str}"
        return False, error_msg
    
    return False, f"{field_name}: Unknown comparison type"

def compare_generic(ml_value, external_value, comparison_type="PCOM", field_name=""):
    """
    Generic comparison for other fields.
    Returns: (is_match, error_message)
    """
    if not ml_value or ml_value == "Not Found":
        return False, f"{field_name}: Missing in Machine Log"
    
    if not external_value or external_value == "Not Found":
        return False, f"{field_name}: Missing in {comparison_type}"
    
    ml_str = str(ml_value).strip().upper()
    external_str = str(external_value).strip().upper()
    
    print(f"\n🔍 COMPARING {field_name} ({comparison_type}):")
    print(f"   ML: {ml_str}")
    print(f"   External: {external_str}")
    
    if ml_str == external_str:
        print(f"   ✅ Match")
        return True, ""
    
    error_msg = f"{field_name}: Mismatch in {comparison_type} - Expected: {ml_str}, Got: {external_str}"
    return False, error_msg

# ========== MAIN VALIDATION FUNCTION ==========
def main_airtel(filepath, pcom_path=None, cnum_path=None, sim_oda_path=None, image_paths=None):
    """
    Main AIRTEL validation function - UPDATED: Includes image handling
    """
    print("="*80)
    print("🚀 AIRTEL VALIDATION STARTED (with 80% cps similarity check)")
    print(f"📸 Image paths received: {image_paths}")
    print("="*80)
    
    validation_errors = []
    detailed_errors = []  # For detailed error messages to show in UI
    
    try:
        # Parse Machine Log first
        print("\n" + "="*80)
        machine_fields = parse_machine_log(filepath)
        
        # Get non-"Not Found" ML values for cps search
        ml_values_for_search = {}
        for field_name, ml_value in machine_fields.items():
            if ml_value != "Not Found" and ml_value:
                ml_values_for_search[field_name] = ml_value
        
        print(f"\n📋 Machine Log values for cps search: {len(ml_values_for_search)} fields")
        
        # Parse other files
        pcom_fields = parse_pcom_file(pcom_path) if pcom_path else {}
        cnum_fields = parse_cnum_file(cnum_path) if cnum_path else {}
        
        # Parse cps - UPDATED with 80% similarity check
        cps_fields = {}
        if sim_oda_path and os.path.exists(sim_oda_path):
            cps_fields = parse_cps_file(sim_oda_path, ml_values_for_search)
        else:
            print("⚠️  cps file not provided or not found")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Validation Report"
        
        # Setup styles and headers with updated template
        styles = setup_excel_styles()
        header_row = setup_excel_headers(ws, styles, filepath, pcom_path, cnum_path, sim_oda_path)
        
        # Add image section BEFORE validation table if images are provided
        current_row = header_row
        
        if image_paths and any(image_paths.values()):
            print("\n📸 Adding images to report...")
            # Add an empty row for spacing
            ws.row_dimensions[current_row].height = 10
            current_row += 1
            
            # Add image section
            current_row = add_image_section_to_report(ws, styles, image_paths, current_row)
        
        # Headers for validation table - NOW at the current_row position
        headers = [
            "Field", "Machine Log", "PCOM", "CNUM", "CPS File",
            "PCOM Status", "CNUM Status", "CPS File Status", "Validation Status"
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.fill = styles['dark_blue_fill']
            cell.font = styles['header_font']
            cell.border = styles['thick_border']
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Fill data rows
        row = current_row + 1
        
        for field_name, rules in AIR_TEL_VALIDATION_RULES.items():
            ml_rule, pcom_rule, cnum_rule, cps_rule = rules
            
            # Get values from Machine Log
            ml_value = machine_fields.get(field_name, "Not Found")
            
            # For PCOM column
            if pcom_rule == "NR":
                pcom_display = "NR"
                pcom_actual_value = pcom_fields.get(field_name, "")
            elif pcom_path:
                pcom_actual_value = pcom_fields.get(field_name, "Not Found")
                pcom_display = pcom_actual_value
            else:
                pcom_display = "N/A"
                pcom_actual_value = ""
            
            # For CNUM column
            if cnum_rule == "NR":
                cnum_display = "NR"
                cnum_actual_value = cnum_fields.get(field_name, "")
            elif cnum_path:
                cnum_actual_value = cnum_fields.get(field_name, "Not Found")
                cnum_display = cnum_actual_value
            else:
                cnum_display = "N/A"
                cnum_actual_value = ""
            
            # For cps column - Only show values that passed 80% similarity check
            if cps_rule == "NR":
                cps_display = "NR"
                cps_actual_value = ""
            elif sim_oda_path:
                # Get value that passed 80% similarity check
                cps_actual_value = cps_fields.get(field_name, "Not Found")
                cps_display = cps_actual_value
            else:
                cps_display = "N/A"
                cps_actual_value = ""
            
            # Initialize statuses
            pcom_status = "NR" if pcom_rule == "NR" else "N/A"
            cnum_status = "NR" if cnum_rule == "NR" else "N/A"
            cps_status = "NR" if cps_rule == "NR" else "N/A"
            overall_status = "✅ Pass"
            
            # Check Machine Log
            if ml_rule == "from_value" and ml_value == "Not Found":
                overall_status = "❌ Fail"
                error_msg = f"{field_name}: Missing in Machine Log"
                validation_errors.append(error_msg)
                detailed_errors.append(error_msg)
            
            # Check PCOM validation
            if pcom_rule == "from_value" and pcom_path:
                if pcom_actual_value == "Not Found":
                    pcom_status = "❌ Fail"
                    overall_status = "❌ Fail"
                    error_msg = f"{field_name}: Missing in PCOM"
                    validation_errors.append(error_msg)
                    detailed_errors.append(error_msg)
                else:
                    # Use appropriate comparison function
                    if field_name in ["PUK1", "PUK2"]:
                        is_match, error_msg = compare_puk(ml_value, pcom_actual_value, "PCOM", field_name)
                    elif field_name == "ICCID":
                        is_match, error_msg = compare_iccid(ml_value, pcom_actual_value, "PCOM", field_name)
                    elif field_name == "IMSI":
                        is_match, error_msg = compare_imsi(ml_value, pcom_actual_value, "PCOM", field_name)
                    else:
                        is_match, error_msg = compare_generic(ml_value, pcom_actual_value, "PCOM", field_name)
                    
                    if not is_match:
                        pcom_status = "❌ Fail"
                        overall_status = "❌ Fail"
                        validation_errors.append(error_msg)
                        detailed_errors.append(error_msg)
                    else:
                        pcom_status = "✅ Pass"
            
            # Check CNUM validation
            if cnum_rule == "from_value" and cnum_path:
                if cnum_actual_value == "Not Found":
                    cnum_status = "❌ Fail"
                    overall_status = "❌ Fail"
                    error_msg = f"{field_name}: Missing in CNUM"
                    validation_errors.append(error_msg)
                    detailed_errors.append(error_msg)
                else:
                    # Use appropriate comparison function
                    if field_name in ["PUK1", "PUK2"]:
                        is_match, error_msg = compare_puk(ml_value, cnum_actual_value, "CNUM", field_name)
                    elif field_name == "ICCID":
                        is_match, error_msg = compare_iccid(ml_value, cnum_actual_value, "CNUM", field_name)
                    elif field_name == "IMSI":
                        is_match, error_msg = compare_imsi(ml_value, cnum_actual_value, "CNUM", field_name)
                    else:
                        is_match, error_msg = compare_generic(ml_value, cnum_actual_value, "CNUM", field_name)
                    
                    if not is_match:
                        cnum_status = "❌ Fail"
                        overall_status = "❌ Fail"
                        validation_errors.append(error_msg)
                        detailed_errors.append(error_msg)
                    else:
                        cnum_status = "✅ Pass"
            
            # Check cps validation - Only validate if we have a value that passed 80% check
            if cps_rule == "from_value" and sim_oda_path:
                if cps_actual_value == "Not Found":
                    # Value didn't pass 80% similarity check
                    cps_status = "❌ Fail"
                    overall_status = "❌ Fail"
                    error_msg = f"{field_name}: Not found in cps (or below 80% similarity)"
                    validation_errors.append(error_msg)
                    detailed_errors.append(error_msg)
                else:
                    # We have a value that passed 80% similarity, now compare for exact match
                    if field_name == "ICCID":
                        is_match, error_msg = compare_iccid(ml_value, cps_actual_value, "cps", field_name)
                    elif field_name == "IMSI":
                        is_match, error_msg = compare_imsi(ml_value, cps_actual_value, "cps", field_name)
                    else:
                        is_match, error_msg = compare_generic(ml_value, cps_actual_value, "cps", field_name)
                    
                    if not is_match:
                        cps_status = "❌ Fail"
                        overall_status = "❌ Fail"
                        validation_errors.append(error_msg)
                        detailed_errors.append(error_msg)
                    else:
                        cps_status = "✅ Pass"
            
            # Write row with UPDATED styling
            data = [
                field_name,
                ml_value,
                pcom_display,
                cnum_display,
                cps_display,
                pcom_status,
                cnum_status,
                cps_status,
                overall_status
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=str(value))
                cell.border = styles['thick_border']
                cell.number_format = '@'
                
                cell_value_str = str(value).strip()
                
                # Apply exact color coding from template
                if "✅ Pass" in cell_value_str:
                    cell.fill = styles['green_fill']  # #BDDDE9
                elif "❌ Fail" in cell_value_str:
                    cell.fill = styles['red_fill']    # #FF0000
                elif "Not Found" in cell_value_str and col > 1:  # Only color data cells, not field names
                    cell.fill = styles['yellow_fill'] # #FFFF00
                elif cell_value_str == "NR":
                    cell.fill = styles['yellow_fill']   # Using yellow for NR (was gray)
                elif cell_value_str == "N/A":
                    cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
            
            row += 1
        
        # Add detailed errors section
        if detailed_errors:
            error_row = row + 2
            ws.cell(row=error_row, column=1, value="DETAILED ERROR ANALYSIS").font = Font(size=14, bold=True)
            ws.merge_cells(f'A{error_row}:L{error_row}')
            error_row += 1
            
            for i, error in enumerate(detailed_errors, 1): 
                ws.cell(row=error_row, column=1, value=f"{i}. {error}").font = Font(color='FF0000')
                error_row += 1
        
        # Auto-adjust columns with SPECIAL HANDLING for Column A (50% wider)
        for col in range(1, len(headers) + 1):
            max_length = 0
            col_letter = get_column_letter(col)
            
            for r in range(current_row, row + 1):
                cell_value = ws.cell(row=r, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            # SPECIAL: Increase Column A width by 50%
            if col == 1:  # Column A
                base_width = min(max_length + 4, 50)
                adjusted_width = base_width * 1.5  # Increase by 50%
                ws.column_dimensions[col_letter].width = adjusted_width
            else:
                adjusted_width = min(max_length + 4, 50)
                ws.column_dimensions[col_letter].width = adjusted_width
        
        # Save report using ICCID naming
        # Extract ICCID for naming
        iccid = machine_fields.get("ICCID")
        if not iccid or iccid == "Not Found":
            # Try parsing ICCID from cps if machine log fails
            iccid = cps_fields.get("ICCID")
        
        # Extract SOF number from folder path for report naming
        folder_path = os.path.dirname(filepath) if filepath else ""
        sof_number = extract_sof_number(folder_path)
        
        report_path = save_report(wb, filepath, pcom_path, iccid, sof_number)
        print(f"\n✅ Validation report saved: {report_path}")
        
        # Summary - Include image information
        ml_found = sum(1 for v in machine_fields.values() if v != "Not Found")
        cps_found = len([v for v in cps_fields.values() if v != "Not Found"])
        
        image_count = len([path for path in image_paths.values() if path and os.path.exists(path)]) if image_paths else 0
        print(f"\n📊 SUMMARY:")
        print(f"   • Machine Log fields: {ml_found}/{len(AIR_TEL_VALIDATION_RULES)}")
        print(f"   • cps valid values (≥80% similarity): {cps_found} fields")
        print(f"   • Images provided: {image_count}")
        print(f"   • Validation errors: {len(validation_errors)}")
        
        if detailed_errors:
            print(f"\n❌ DETAILED ERRORS:")
            for i, error in enumerate(detailed_errors[:20]):
                print(f"   {i+1}. {error}")
        
        return report_path, detailed_errors
        
    except Exception as e:
        error_msg = f"Validation error: {str(e)}"
        print(f"\n❌ {error_msg}")
        traceback.print_exc()
        return None, [error_msg]

# ========== COMPATIBILITY FUNCTIONS ==========
def run_airtel_validation(filepath, pcom_path=None, cnum_path=None, cps_path=None, image_paths=None):
    """
    Compatibility function that matches your GUI's expected call
    Returns: (report_path, detailed_errors)
    """
    return main_airtel(filepath, pcom_path, cnum_path, cps_path, image_paths)

if __name__ == "__main__":
    # Test the validation
    print("AIRTEL Validation Module - With 80% cps Similarity Check and Image Support")
    print("="*80)
    
    test_choice = input("\nRun validation test? (y/n): ").strip().lower()
    if test_choice == 'y':
        test_ml = input("Enter test Machine Log path: ").strip()
        
        if not os.path.exists(test_ml):
            print("❌ File not found!")
        else:
            pcom = input("Enter PCOM path (optional): ").strip()
            cnum = input("Enter CNUM path (optional): ").strip()
            cps = input("Enter cps path (optional): ").strip()
            
            if pcom and not os.path.exists(pcom):
                print("⚠️  PCOM not found, skipping...")
                pcom = None
            
            if cnum and not os.path.exists(cnum):
                print("⚠️  CNUM not found, skipping...")
                cnum = None
            
            if cps and not os.path.exists(cps):
                print("⚠️  cps not found, skipping...")
                cps = None
            
            print("\n" + "="*80)
            print("🚀 STARTING VALIDATION...")
            print("="*80)
            
            report, errors = run_airtel_validation(test_ml, pcom, cnum, cps)
            
            if report:
                print(f"\n✅ Validation completed! Report: {report}")
                if errors:
                    print(f"\n❌ Errors found: {len(errors)}")
                    for i, error in enumerate(errors, 1):
                        print(f"   {i}. {error}")
                if os.name == 'nt':
                    os.startfile(report)
            else:
                print("\n❌ Validation failed!")